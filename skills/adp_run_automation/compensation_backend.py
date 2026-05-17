#!/usr/bin/env python3
"""skills/adp_run_automation/compensation_backend - Playwright-driven Earnings & Hours XLSX extractor.

Source: ADP RUN > Reports > My saved custom reports > "Earnings and Hours V1"
(the saved custom report; see selectors/compensation.json for portability
caveats). Renders an in-page modal, queues an Excel export, and downloads to
extracted/downloads/Earnings-and-Hours-V1.xlsx.

Mirrors skills/adp_run_automation/shift_backend.py:
    * `build_plan()` -> Playwright playbook for the AI agent (login + nav +
      filters + download).
    * `parse_xlsx()` -> per-line records (one per earning line on a paycheck).
    * `infer_wage_rates()` -> one record per employee with their canonical
      wage rate plus salaried/multi-rate flags.
    * `compensation()` -> high-level public entry that combines parse + infer
      and returns the shape promised in the bhaga-daily-refresh plan:
      [{employee_id, employee_name, wage_rate_dollars, is_salaried}, ...]

Critical gotchas (see selectors/compensation.json for full notes):
    * The custom report is per-store; "Earnings and Hours V1" exists in
      Palmetto's account. Other stores must create the equivalent report
      first (creation flow not yet documented).
    * Employee Name column is sparse -- populated only on the first row of
      each employee block; forward-fill required.
    * Hours are DECIMAL (e.g. 30.18 == 30h 11m), NOT 'H:MM' like the
      Timecard XLSX. Don't reuse parse_hhmm_to_decimal here.
    * Voided/reissued paychecks emit reversal triplets ($X, -$X, $X) on
      the same check date. Wage-rate inference ignores them by filtering
      to Hours>0 Regular rows.
    * Earnings report uses 'LastName, FirstName' format whereas the
      Timecard report uses 'LastName FirstName' (no comma). The aliases
      map must handle both forms for cross-backend joins.
    * Lindsay Krause is paid HOURLY at $25/hr (not salaried). Exclusion
      from labor% / tip pool is by NAME via the store profile, not by
      the is_salaried flag.

Status (2026-05-16): proven from-scratch end-to-end on Palmetto Superfoods.
Date range Mar 22 - May 15 2026, 12 employees, 116 data rows.
"""

from __future__ import annotations

import datetime
import json
import os
import pathlib
import subprocess
import sys
from typing import Optional

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))

from core.config_loader import project_dir
from skills.adp_run_automation.shift_backend import normalize_employee_name
from skills.credentials import registry as cred_registry


_PROJECT = pathlib.Path(project_dir())
DOWNLOADS_DIR = _PROJECT / "extracted" / "downloads"
SELECTORS_PATH = _PROJECT / "skills" / "adp_run_automation" / "selectors" / "compensation.json"

LOGIN_URL = "https://runpayroll.adp.com"

# Default name of the saved custom report. Read this from the store profile
# (`adp_wage_rate_report_name`) in production; this is the Palmetto default.
DEFAULT_SAVED_REPORT_NAME = "Earnings and Hours V1"


# ── Credentials (mirrors shift_backend) ───────────────────────────


def _credential_name(store: str) -> str:
    return f"adp_{store.lower()}_login"


def get_credentials(store: str = "palmetto") -> dict:
    entry = cred_registry.lookup(_credential_name(store))
    if not entry:
        raise RuntimeError(
            f"No credential '{_credential_name(store)}' in registry. "
            f"Register via skills/browser/collaborative.py or the credentials skill."
        )
    if entry.get("type") != "keychain":
        raise RuntimeError(
            f"Credential '{_credential_name(store)}' is type {entry.get('type')!r}, "
            f"expected 'keychain'."
        )
    result = subprocess.run(
        [
            "security", "find-generic-password",
            "-a", entry["account"],
            "-s", entry["service"],
            "-w",
        ],
        capture_output=True, text=True, timeout=5,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Keychain lookup failed for {entry['account']}@{entry['service']}: "
            f"{result.stderr.strip()}"
        )
    return {"username": entry["account"], "password": result.stdout.strip()}


# ── Selectors ─────────────────────────────────────────────────────


def selectors() -> dict:
    return json.loads(SELECTORS_PATH.read_text())


# ── XLSX parsing ──────────────────────────────────────────────────


_HEADER_COLUMNS = [
    "Employee Name",
    "Payroll Check Date",
    "Period Start Date",
    "Period End Date",
    "Payroll Earning Hours",
    "Payroll Hourly Earning Rate",
    "Payroll Earning Amount",
    "Payroll Earning Description",
]

# Rows 1-5 are metadata (title, company, IID, date range, generated timestamp);
# row 6 is the column header; data starts at row 7.
_HEADER_ROW = 6


def _coerce_date(v) -> Optional[datetime.date]:
    """openpyxl returns datetime; sometimes strings sneak through. Be defensive."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_xlsx(
    xlsx_path: pathlib.Path,
    *,
    employee_aliases: Optional[dict] = None,
) -> list[dict]:
    """Parse the Earnings & Hours XLSX into per-line earning records.

    One record per row in the report (so multiple per paycheck and many per
    employee). Employee Name is forward-filled from the sparse source format.

        {
            "employee_name": str,        # forward-filled, alias-normalized
            "raw_employee_name": str,    # original (only on the first row of each block, else copied via fill)
            "check_date": "YYYY-MM-DD",
            "period_start": "YYYY-MM-DD",
            "period_end": "YYYY-MM-DD",
            "hours": float | None,       # decimal hours (NOT H:MM)
            "hourly_rate": float | None, # dollars/hour
            "amount": float,             # dollars (can be negative for reversals)
            "description": str,          # e.g. 'Regular', 'Overtime', 'Cash tips', ...
        }

    Records preserve source order (which is sorted by employee, then check date).
    Requires openpyxl.
    """
    import openpyxl

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[_HEADER_ROW]]
    if header[: len(_HEADER_COLUMNS)] != _HEADER_COLUMNS:
        raise ValueError(
            f"Earnings xlsx header mismatch at row {_HEADER_ROW}. "
            f"Expected {_HEADER_COLUMNS}, got {header[: len(_HEADER_COLUMNS)]}. "
            f"ADP may have changed the saved-report column layout, or the wrong "
            f"saved report was downloaded."
        )

    records: list[dict] = []
    current_raw_name: Optional[str] = None

    for raw_row in ws.iter_rows(min_row=_HEADER_ROW + 1, values_only=True):
        row = dict(zip(header, raw_row))

        raw_name = row.get("Employee Name")
        if raw_name:
            current_raw_name = str(raw_name).strip()
        if not current_raw_name:
            # Stray row before any employee block (shouldn't happen with valid ADP export).
            continue

        check_date = _coerce_date(row.get("Payroll Check Date"))
        amount = _coerce_float(row.get("Payroll Earning Amount"))
        if check_date is None or amount is None:
            continue

        records.append({
            "employee_name": normalize_employee_name(current_raw_name, employee_aliases),
            "raw_employee_name": current_raw_name,
            "check_date": check_date.isoformat(),
            "period_start": (
                _coerce_date(row.get("Period Start Date")).isoformat()
                if _coerce_date(row.get("Period Start Date")) else None
            ),
            "period_end": (
                _coerce_date(row.get("Period End Date")).isoformat()
                if _coerce_date(row.get("Period End Date")) else None
            ),
            "hours": _coerce_float(row.get("Payroll Earning Hours")),
            "hourly_rate": _coerce_float(row.get("Payroll Hourly Earning Rate")),
            "amount": amount,
            "description": str(row.get("Payroll Earning Description") or "").strip(),
        })

    return records


# ── Wage-rate inference ───────────────────────────────────────────


def infer_wage_rates(
    earnings: list[dict],
    *,
    excluded_employees: Optional[list[str]] = None,
) -> list[dict]:
    """Collapse per-line earnings into one wage-rate record per employee.

    Inference rules (see selectors/compensation.json#wage_rate_inference_rules):
        1. Filter to Description in {'Regular'} with hours>0 and rate>0.
        2. Group by employee.
        3. Most recent rate (by check_date desc) wins as wage_rate_dollars.
        4. If multiple distinct rates, set multi_rate=True + include rate_history.
        5. If an employee has earnings but ZERO qualifying Regular rows,
           mark is_salaried=True with wage_rate_dollars=None.

    Returns the shape promised in the bhaga-daily-refresh plan:

        [
          {
            "employee_id": str,        # normalized name (until ADP id is wired)
            "employee_name": str,
            "wage_rate_dollars": float | None,
            "ot_rate_dollars": float | None,  # if any Overtime rows seen
            "is_salaried": bool,
            "multi_rate": bool,
            "rate_history": [{"check_date": "YYYY-MM-DD", "rate": float}, ...],
            "excluded_from_labor_pct": bool,  # per store profile excluded_employees
            "raw_employee_names": [str, ...], # original spellings seen (audit aid)
          },
          ...
        ]
    """
    excluded = set(excluded_employees or [])

    by_emp: dict[str, dict] = {}
    for r in earnings:
        emp = r["employee_name"]
        if not emp:
            continue
        bucket = by_emp.setdefault(emp, {
            "employee_id": emp,
            "employee_name": emp,
            "wage_rate_dollars": None,
            "ot_rate_dollars": None,
            "is_salaried": False,
            "multi_rate": False,
            "rate_history": [],
            "ot_rate_history": [],
            "excluded_from_labor_pct": emp in excluded,
            "raw_employee_names": set(),
            "_has_any_earnings": False,
        })
        bucket["raw_employee_names"].add(r["raw_employee_name"])
        bucket["_has_any_earnings"] = True

        if (r["description"] == "Regular"
                and r["hours"] and r["hours"] > 0
                and r["hourly_rate"] and r["hourly_rate"] > 0):
            bucket["rate_history"].append({
                "check_date": r["check_date"],
                "rate": r["hourly_rate"],
            })
        elif (r["description"] == "Overtime"
                and r["hours"] and r["hours"] > 0
                and r["hourly_rate"] and r["hourly_rate"] > 0):
            bucket["ot_rate_history"].append({
                "check_date": r["check_date"],
                "rate": r["hourly_rate"],
            })

    out: list[dict] = []
    for emp, bucket in by_emp.items():
        rh = sorted(bucket["rate_history"], key=lambda x: x["check_date"], reverse=True)
        oh = sorted(bucket["ot_rate_history"], key=lambda x: x["check_date"], reverse=True)
        if rh:
            bucket["wage_rate_dollars"] = rh[0]["rate"]
            distinct = {x["rate"] for x in rh}
            bucket["multi_rate"] = len(distinct) > 1
        else:
            # No qualifying Regular hourly rows -> infer salaried (only if they
            # had any earnings at all; otherwise they're not really on payroll
            # in the window).
            bucket["is_salaried"] = bucket["_has_any_earnings"]
        if oh:
            bucket["ot_rate_dollars"] = oh[0]["rate"]

        bucket["rate_history"] = rh
        bucket["ot_rate_history"] = oh
        bucket["raw_employee_names"] = sorted(bucket["raw_employee_names"])
        bucket.pop("_has_any_earnings", None)
        out.append(bucket)

    out.sort(key=lambda b: b["employee_name"])
    return out


# ── Playwright playbook ───────────────────────────────────────────


def build_plan(
    start_date: datetime.date,
    end_date: datetime.date,
    *,
    store: str = "palmetto",
    saved_report_name: str = DEFAULT_SAVED_REPORT_NAME,
    creds: Optional[dict] = None,
) -> dict:
    """Deterministic Playwright playbook for the AI to execute via user-playwright.

    Flow:
        1. Navigate to LOGIN_URL (cookie session may skip username).
        2. Type password -> Enter.
        3. Click Reports in top nav. Scroll to 'My saved custom reports'.
        4. Click the saved report (e.g. 'Earnings and Hours V1').
        5. In the Custom report builder modal:
            - Ensure date-range preset = 'Custom date range'.
            - Fill From / To textboxes with MM/DD/YYYY.
            - Leave Employment status = Active, People = All, name unchanged.
        6. Click 'Preview report' (data-test-id='view-custom-report').
        7. Click 'Download' button -> menu opens -> click 'Excel (.xlsx)'.
        8. Wait for 'Your report is ready to download' dialog.
        9. Click 'Download report' (data-test-id='download-report').
        10. File saves to extracted/downloads/Earnings-and-Hours-V1.xlsx.
    """
    sels = selectors()
    creds = creds or get_credentials(store)

    return {
        "store": store.lower(),
        "saved_report_name": saved_report_name,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "creds_username": creds["username"],
        "captures": {
            "xlsx_path": None,
            "errors": [],
        },
        "steps": [
            {
                "id": "navigate_login",
                "action": "browser_navigate",
                "args": {"url": LOGIN_URL},
                "description": (
                    "Open ADP RUN login. Cookies may shortcut straight to password."
                ),
            },
            {
                "id": "login_password",
                "action": "browser_type",
                "selectors_hint": "getByRole('textbox', { name: 'Password' })",
                "value_ref": "creds.password",
                "args": {"submit": True},
                "description": (
                    f"Type password and Enter. Keychain handle "
                    f"'{sels['login']['credential_handle']}'."
                ),
                "postcondition": "URL matches /runpayrollmain.adp.com/.*/v2/",
            },
            {
                "id": "wait_for_dashboard",
                "action": "browser_wait_for",
                "args": {"time": 4},
                "description": "Allow SPA to mount.",
            },
            {
                "id": "open_reports",
                "action": "browser_click",
                "selectors_hint": "Top nav: data-test-id='Reports-btn' OR getByRole('button',{name:'Reports'})",
                "description": "Opens the Reports landing page.",
            },
            {
                "id": "wait_for_reports_page",
                "action": "browser_wait_for",
                "args": {"time": 3},
                "description": "Reports page renders categories + My saved custom reports section.",
            },
            {
                "id": "scroll_to_saved_reports",
                "action": "browser_scroll",
                "args": {"to": "My saved custom reports section"},
                "description": (
                    "The saved-reports list is below the system categories. Scroll until visible "
                    "OR use page.locator('text=My saved custom reports').scrollIntoViewIfNeeded()."
                ),
            },
            {
                "id": "open_saved_report",
                "action": "browser_click",
                "selectors_hint": (
                    f"Link/row with exact text '{saved_report_name}' under 'My saved custom reports'."
                ),
                "description": (
                    f"Opens the Custom report builder dialog pre-loaded with '{saved_report_name}'. "
                    f"Filters may show prior values -- always overwrite explicitly."
                ),
            },
            {
                "id": "wait_for_builder_modal",
                "action": "browser_wait_for",
                "args": {"time": 3},
                "description": "Wait for the dialog heading 'Custom report builder' to render with form fields.",
            },
            {
                "id": "ensure_custom_date_range",
                "action": "browser_click_then_option",
                "click_hint": "getByRole('button', { name: /Date range/ }) (combobox-like button)",
                "option_hint": "menuitem/option with name 'Custom date range'",
                "skip_if": "Combobox already displays 'Custom date range'.",
                "description": "Switch the date preset to Custom so From/To textboxes appear.",
            },
            {
                "id": "fill_from_date",
                "action": "browser_fill",
                "selectors_hint": "getByRole('textbox', { name: 'From *' })",
                "args": {"value": start_date.strftime("%m/%d/%Y")},
                "description": "Set From date in MM/DD/YYYY.",
            },
            {
                "id": "fill_to_date",
                "action": "browser_fill",
                "selectors_hint": "getByRole('textbox', { name: 'To *' })",
                "args": {"value": end_date.strftime("%m/%d/%Y")},
                "description": "Set To date in MM/DD/YYYY. Max range is 12 months (ADP-enforced).",
            },
            {
                "id": "click_preview_report",
                "action": "browser_click",
                "selectors_hint": "[data-test-id='view-custom-report'] (aria='Preview report')",
                "description": "Renders the truncated preview grid. Required step before Download appears.",
            },
            {
                "id": "wait_for_preview",
                "action": "browser_wait_for",
                "args": {"time": 4},
                "description": "Wait for the preview grid + Download button to appear.",
            },
            {
                "id": "open_download_menu",
                "action": "browser_click",
                "selectors_hint": "getByRole('button', { name: 'Download' })",
                "description": "Expands the Download dropdown (Excel/PDF menu items).",
            },
            {
                "id": "click_excel_menuitem",
                "action": "browser_click",
                "selectors_hint": "getByRole('menuitem', { name: 'Excel (.xlsx)' })",
                "description": (
                    "ADP queues report generation server-side (~3-10s). A 'Your report is ready "
                    "to download' confirmation dialog will appear when generation completes."
                ),
            },
            {
                "id": "wait_for_ready_dialog",
                "action": "browser_wait_for",
                "args": {"text": "Your report is ready to download", "time": 15},
                "description": "Wait up to 15s for ADP to finish generating the file.",
            },
            {
                "id": "click_download_report",
                "action": "browser_click",
                "selectors_hint": "[data-test-id='download-report']",
                "description": (
                    "Triggers the actual file save. Browser MCP writes to "
                    "extracted/downloads/Earnings-and-Hours-V1.xlsx (spaces -> dashes)."
                ),
            },
            {
                "id": "wait_for_download_flush",
                "action": "browser_wait_for",
                "args": {"time": 3},
                "description": "Let the file flush to disk before parse_xlsx() reads it.",
            },
            {
                "id": "parse_and_validate",
                "action": "python",
                "description": (
                    "earnings = compensation_backend.parse_xlsx(captures.xlsx_path, employee_aliases=...). "
                    "rates = compensation_backend.infer_wage_rates(earnings, excluded_employees=store_profile['excluded']). "
                    "Validate: len(rates) within expected roster size (alert if it drops > 25% from prior run); "
                    "alert via slack on any multi_rate=True (rare; means an employee got a raise mid-window)."
                ),
            },
        ],
    }


# ── Public entry point ────────────────────────────────────────────


def compensation(
    *,
    store: str = "palmetto",
    employee_aliases: Optional[dict] = None,
    excluded_employees: Optional[list[str]] = None,
) -> list[dict]:
    """High-level entry: parse the most recent Earnings & Hours XLSX and return
    one wage-rate record per employee.

    Returns the shape promised in the bhaga-daily-refresh plan
    (with additional audit fields beyond the minimum required):
        [{employee_id, employee_name, wage_rate_dollars, is_salaried,
          ot_rate_dollars, multi_rate, rate_history, excluded_from_labor_pct,
          raw_employee_names}, ...]
    """
    candidates = sorted(
        list(DOWNLOADS_DIR.glob("Earnings*.xlsx"))
        + list(DOWNLOADS_DIR.glob("Earnings-and-Hours*.xlsx")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    # dedupe by path
    seen = set()
    deduped = []
    for p in candidates:
        if p not in seen:
            deduped.append(p)
            seen.add(p)
    if not deduped:
        raise FileNotFoundError(
            f"No Earnings*.xlsx found in {DOWNLOADS_DIR}. Run build_plan() and "
            f"have the AI drive Playwright to populate."
        )

    earnings = parse_xlsx(deduped[0], employee_aliases=employee_aliases)
    return infer_wage_rates(earnings, excluded_employees=excluded_employees)


def raw_earnings(
    *,
    store: str = "palmetto",
    employee_aliases: Optional[dict] = None,
) -> list[dict]:
    """Per-line earnings records from the most recent download. Useful for
    debugging or for downstream tip reconciliation (the same xlsx contains
    'Credit Card Tips Owed' and 'Misc reimbursement non-taxable' lines that
    we can cross-check against Square tips)."""
    candidates = sorted(
        list(DOWNLOADS_DIR.glob("Earnings*.xlsx")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"No Earnings*.xlsx found in {DOWNLOADS_DIR}.")
    return parse_xlsx(candidates[0], employee_aliases=employee_aliases)


# ── CLI ───────────────────────────────────────────────────────────


if __name__ == "__main__":
    import argparse

    cli = argparse.ArgumentParser(description="ADP RUN Earnings & Hours report extractor")
    sub = cli.add_subparsers(dest="cmd")

    p_parse = sub.add_parser("parse", help="Parse an existing Earnings xlsx.")
    p_parse.add_argument("xlsx_path")
    p_parse.add_argument("--aliases", default=None,
        help="JSON file with {raw_name: canonical_name} alias map.")
    p_parse.add_argument("--rates", action="store_true",
        help="Run infer_wage_rates() instead of returning per-line earnings.")
    p_parse.add_argument("--excluded", default=None,
        help="JSON list of canonical employee names to mark excluded_from_labor_pct.")

    p_comp = sub.add_parser("compensation", help="Parse most recent download and infer wage rates.")
    p_comp.add_argument("--store", default="palmetto")
    p_comp.add_argument("--aliases", default=None)
    p_comp.add_argument("--excluded", default=None)

    p_plan = sub.add_parser("plan", help="Print the Playwright playbook for a date range.")
    p_plan.add_argument("--start", required=True, help="YYYY-MM-DD")
    p_plan.add_argument("--end", required=True, help="YYYY-MM-DD")
    p_plan.add_argument("--store", default="palmetto")
    p_plan.add_argument("--report-name", default=DEFAULT_SAVED_REPORT_NAME)

    p_creds = sub.add_parser("verify-creds", help="Check Keychain access.")
    p_creds.add_argument("--store", default="palmetto")

    args = cli.parse_args()
    aliases = json.loads(pathlib.Path(args.aliases).read_text()) if getattr(args, "aliases", None) else None
    excluded = json.loads(pathlib.Path(args.excluded).read_text()) if getattr(args, "excluded", None) else None

    if args.cmd == "parse":
        earnings = parse_xlsx(pathlib.Path(args.xlsx_path), employee_aliases=aliases)
        if args.rates:
            print(json.dumps(infer_wage_rates(earnings, excluded_employees=excluded), indent=2))
        else:
            print(json.dumps(earnings, indent=2))
    elif args.cmd == "compensation":
        print(json.dumps(
            compensation(store=args.store, employee_aliases=aliases, excluded_employees=excluded),
            indent=2,
        ))
    elif args.cmd == "plan":
        plan = build_plan(
            datetime.date.fromisoformat(args.start),
            datetime.date.fromisoformat(args.end),
            store=args.store,
            saved_report_name=args.report_name,
        )
        print(json.dumps(plan, indent=2))
    elif args.cmd == "verify-creds":
        creds = get_credentials(args.store)
        print(json.dumps({
            "store": args.store,
            "username": creds["username"],
            "password_length": len(creds["password"]),
            "password_preview": creds["password"][:2] + "***" + creds["password"][-2:],
        }, indent=2))
    else:
        cli.print_help()
