"""Standalone Playwright runner for ADP RUN report downloads.

Two scrapes:
    download_timecard(...)  — Reports > Time reports > Timecard, exports .xlsx
                              of all employee punches across selected pay periods.
                              Run nightly.
    download_earnings(...)  — Reports > My saved custom reports > "Earnings and
                              Hours V1", exports .xlsx with wage rates +
                              Credit-Card-Tips-Owed lines per paycheck.
                              Only needed every 14 days (after each pay-period
                              close), but cheap to run nightly.
    download_schedule(...)  — Home > Team Schedule ("Manage Schedules" grid).
                              Scrapes per-day SCHEDULED labor hours for the
                              current + next week directly from the grid DOM
                              (no file export exists). See schedule_backend.py.
                              Run nightly; forward-looking, no target_date.

ADP quirks accounted for:
    - Timecard report runs INSIDE iframe[name="mdfTimeFrame"] -- all selectors
      must traverse the iframe.
    - Export-to-Excel is a custom <SDF-BUTTON> with id "report-excel-button";
      role-based selectors don't always find it, so we fall back to JS click.
    - Earnings report uses an async generation flow: Download dropdown -> Excel
      menu item -> wait for "Your report is ready" dialog -> click Download report.
    - Login may or may not be prompted: cached session lands directly on dashboard;
      expired session shows a password prompt (no username; ADP remembers it).
    - If MFA is challenged, we ALERT-AND-EXIT rather than try to drive it (TODO
      M3.10: wire up skills/slack.request_otp for OTP delivery).
"""

from __future__ import annotations

import argparse
import datetime
import json
import math
import os
import pathlib
import re
import subprocess
import sys
import time
from typing import Optional

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))

from skills._browser_runtime.runtime import (
    DOWNLOADS_DIR,
    download_to,
    is_fresh_download,
    launch_persistent,
)
from skills.credentials import registry as cred_registry

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[2]
STORE_PROFILES = PROJECT_ROOT / "agents" / "bhaga" / "knowledge-base" / "store-profiles"

# ADP retired the bare https://runpayroll.adp.com entry point (2026-06-28): it now
# server-redirects to https://sorry.adp.com/sorry/. The live login flow is reachable
# via /enrollment.aspx, which routes through ADP's federation redirector to the
# sign-in SPA (online.adp.com/signin/v1/?APPID=RUN&productId=...) with the correct,
# self-supplied productId. Using enrollment.aspx lets ADP resolve the current
# productId rather than hardcoding a value that can rotate.
LOGIN_URL = "https://runpayroll.adp.com/enrollment.aspx"
POST_LOGIN_URL_RE = re.compile(r"runpayrollmain\.adp\.com/.*/v2/")

# Match ADP's "Current Pay Period" / "Current" / "This Pay Period" dropdown
# entry that covers the IN-FLIGHT payroll window. Closed pay-period entries
# render as date ranges ("05/05/2026 - 05/18/2026") and don't intersect.
# Allow a trailing parenthetical/date suffix (ADP sometimes appends the open
# window). Do NOT match "Current Tax Period" / "Last Pay Period".
# Module-level so unit tests can import without launching Playwright.
_CURRENT_PAY_PERIOD_RE = re.compile(
    r"^(?:current\s+pay\s+period|this\s+pay\s+period)(?:\s*[(\[].*)?$|^current$",
    re.I,
)
# Date-range text inside a Pay Period option label.
_PAY_PERIOD_DATE_RANGE_RE = re.compile(
    r"(\d{1,2})/(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})"
)
# Timecard Excel export can be slow when Select All is the only remaining
# fallback (full history). Single-period exports finish well under this.
_TIMECARD_DOWNLOAD_TIMEOUT_MS = 180_000
KEYCHAIN_SERVICE_TEMPLATE = "jarvis-adp-{store}"


def _biweekly_period_bounds(
    target: datetime.date,
    *,
    anchor_end: datetime.date,
    period_len: int = 14,
) -> tuple[datetime.date, datetime.date]:
    """Return the biweekly [start, end] containing ``target``.

    Same cadence as ``update_model_sheet.discover_periods``: a period ENDS
    on ``anchor_end`` and every ±period_len days from there; each period
    spans ``[end - (period_len - 1), end]`` inclusive.

    Uses ceil so an in-flight day after the latest closed end (e.g. 2026-07-15
    when the closed end is 2026-07-12) maps to the open window ending
    ``anchor + k*period_len`` that still contains ``target``.
    """
    if period_len < 1:
        raise ValueError(f"period_len must be >= 1, got {period_len}")
    k = math.ceil((target - anchor_end).days / period_len)
    end = anchor_end + datetime.timedelta(days=period_len * k)
    start = end - datetime.timedelta(days=period_len - 1)
    return start, end


def _period_len_from_frequency(pay_frequency: str) -> int:
    freq = (pay_frequency or "").strip().lower()
    if freq == "biweekly":
        return 14
    raise ValueError(
        f"unsupported pay_frequency {pay_frequency!r}; only 'Biweekly' is implemented"
    )


def _format_pay_period_label(start: datetime.date, end: datetime.date, *, padded: bool) -> str:
    if padded:
        return f"{start.month:02d}/{start.day:02d}/{start.year} - {end.month:02d}/{end.day:02d}/{end.year}"
    return f"{start.month}/{start.day}/{start.year} - {end.month}/{end.day}/{end.year}"


def _parse_pay_period_range(text: str) -> Optional[tuple[datetime.date, datetime.date]]:
    m = _PAY_PERIOD_DATE_RANGE_RE.search(text or "")
    if not m:
        return None
    start = datetime.date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    end = datetime.date(int(m.group(6)), int(m.group(4)), int(m.group(5)))
    return start, end


def _is_current_pay_period_label(text: str) -> bool:
    return bool(_CURRENT_PAY_PERIOD_RE.match((text or "").strip()))


def _enumerate_pay_period_options(frame) -> list[tuple[object, str]]:
    """Return [(locator, accessible_name), ...] for every Pay Period option."""
    opts = frame.get_by_role("option")
    try:
        n = opts.count()
    except Exception:  # noqa: BLE001
        return []
    out: list[tuple[object, str]] = []
    for i in range(n):
        opt = opts.nth(i)
        try:
            name = (opt.get_attribute("aria-label") or opt.inner_text() or "").strip()
        except Exception:  # noqa: BLE001
            name = ""
        out.append((opt, name))
    return out


def _select_pay_period_for_target(
    frame,
    page,
    *,
    target_date: datetime.date,
    store: str,
) -> None:
    """Click the single Pay Period option that covers ``target_date``.

    Enumerates options in Python (Playwright ``name=`` filters miss ADP's
    custom accessible names). Falls back to Select All only as last resort,
    logging the option names seen for diagnosis.
    """
    options = _enumerate_pay_period_options(frame)
    names = [n for _, n in options if n]

    # 1) Any option whose embedded date range contains target_date.
    for opt, name in options:
        bounds = _parse_pay_period_range(name)
        if not bounds:
            continue
        start, end = bounds
        if start <= target_date <= end:
            opt.click()
            print(
                f"[adp_timecard] selected pay period "
                f"{start.isoformat()} → {end.isoformat()} "
                f"(contains target_date={target_date.isoformat()})"
            )
            return

    # 2) Click by store-profile-derived biweekly bounds (unpadded + padded).
    try:
        profile = _load_store_profile(store)
        adp = profile.get("adp_run") or {}
        anchor = datetime.date.fromisoformat(adp["pay_periods_anchor_end_date"])
        period_len = _period_len_from_frequency(adp.get("pay_frequency", "Biweekly"))
        want_start, want_end = _biweekly_period_bounds(
            target_date, anchor_end=anchor, period_len=period_len
        )
        needles = {
            _format_pay_period_label(want_start, want_end, padded=False),
            _format_pay_period_label(want_start, want_end, padded=True),
        }
        for opt, name in options:
            if any(needle in name for needle in needles):
                opt.click()
                print(
                    f"[adp_timecard] selected pay period via profile bounds "
                    f"{want_start.isoformat()} → {want_end.isoformat()} "
                    f"(target_date={target_date.isoformat()})"
                )
                return
    except Exception as exc:  # noqa: BLE001
        print(f"[adp_timecard] WARN: profile period bounds failed: {exc}")

    # 3) "Current Pay Period" (in-flight), including labels with a date suffix.
    for opt, name in options:
        if _is_current_pay_period_label(name):
            opt.click()
            print(
                f"[adp_timecard] no closed pay period contains "
                f"{target_date.isoformat()}; selected current-period option "
                f"{name!r}"
            )
            return

    # 4) Select All last — large export; download timeout is raised to match.
    print(
        f"[adp_timecard] WARN: no closed/current pay period for "
        f"{target_date.isoformat()}; falling back to Select All. "
        f"options_seen={names!r}"
    )
    try:
        frame.get_by_role("option", name=re.compile(r"^Select All$", re.I)).first.click()
    except Exception:  # noqa: BLE001
        for opt, name in options:
            if re.search(r"^Select All$", name or "", re.I):
                opt.click()
                break
        else:
            print("[adp_timecard] WARN: Select All option not clickable")


# ── Credentials ────────────────────────────────────────────────────


def _get_adp_password(store: str) -> str:
    """Pull ADP password via dual-backend get_secret (Keychain or GCP Secret Manager)."""
    return cred_registry.get_secret(f"adp_{store}_login")


# ── Login (shared by both scrapes) ─────────────────────────────────


def _get_adp_username(store: str) -> str:
    """Read the ADP User ID from the credential registry.

    The registry entry's `account` field holds the user-visible User ID
    (e.g. an email). Same entry used by _get_adp_password() — single source
    of truth so password rotations don't require touching the username.
    """
    cred_name = f"adp_{store}_login"
    entry = cred_registry.lookup(cred_name)
    if not entry or "account" not in entry:
        raise RuntimeError(
            f"Credential '{cred_name}' not found in registry or missing 'account' field. "
            f"Register it with: python -m skills.credentials.registry register ..."
        )
    return entry["account"]


def _load_login_page(page, *, timeout_ms: int = 60_000) -> None:
    """Navigate to ADP login and wait for the page to settle.

    ADP's login SPA can take 15-20+ seconds to hydrate from non-US
    locations (CDN + JS bundle latency). We give it generous timeouts
    and wait for networkidle so the SDF web-component framework has
    time to render the login form.
    """
    print(f"[adp_login] step=goto url={LOGIN_URL}")
    t0 = time.monotonic()
    page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=timeout_ms)
    print(f"[adp_login] step=domcontentloaded "
          f"elapsed={time.monotonic() - t0:.1f}s url={page.url}")
    try:
        page.wait_for_load_state("networkidle", timeout=30_000)
    except Exception:  # noqa: BLE001
        pass
    print(f"[adp_login] step=page-settled "
          f"elapsed={time.monotonic() - t0:.1f}s url={page.url}")


_SORRY_ADP_HOST = "sorry.adp.com"
_LOGIN_RETRY_BACKOFF_BASE_S = 3  # base sleep between login retry attempts


def _is_throttled(page) -> bool:
    """Return True if the current page URL is ADP's throttle interstitial."""
    return _SORRY_ADP_HOST in (page.url or "")


def _wait_for_login_form(page, *, max_retries: int = 2, _sleep_fn=None):
    """Wait for the User ID textbox to appear, recovering on stall or throttle.

    Two failure modes are handled:

    1. **JS-hydration stall** (page.url is still the login SPA but the form
       never renders): re-navigate to LOGIN_URL to trigger a fresh hydration.
    2. **ADP throttle interstitial** (page.url is sorry.adp.com/sorry/):
       ``page.reload()`` is wrong here — it re-requests the sorry URL and
       can never get back to the login SPA. We must issue a fresh
       ``page.goto(LOGIN_URL)`` to escape the throttle, with exponential
       backoff so a transient rate-limit has time to clear.

    On exhaustion: if the final page is still on the throttle host, raises
    ``AdpLoginThrottled`` so daily_refresh can treat it as a graceful ADP skip
    (next nightly retries). Any other stall raises ``RuntimeError`` as before.

    Returns the User ID locator once visible.
    """
    from agents.bhaga.scripts.otp_gate import AdpLoginThrottled  # local import

    uid_box = page.get_by_role("textbox", name=re.compile(r"^User ID$", re.I)).first
    total_attempts = max_retries + 1

    sleep = _sleep_fn if _sleep_fn is not None else time.sleep

    for attempt in range(1, total_attempts + 1):
        print(f"[adp_login] step=wait-uid-box (attempt {attempt}/{total_attempts})")
        t0 = time.monotonic()
        try:
            uid_box.wait_for(state="visible", timeout=60_000)
            print(f"[adp_login] step=uid-box-visible "
                  f"elapsed={time.monotonic() - t0:.1f}s (attempt {attempt}/{total_attempts})")
            return uid_box
        except Exception:  # noqa: BLE001
            elapsed = time.monotonic() - t0
            throttled = _is_throttled(page)
            print(f"[adp_login] step=uid-box-timeout "
                  f"elapsed={elapsed:.1f}s (attempt {attempt}/{total_attempts}) "
                  f"throttled={throttled} url={page.url}")

        if attempt <= max_retries:
            backoff_s = _LOGIN_RETRY_BACKOFF_BASE_S * (2 ** (attempt - 1))
            print(f"[adp_login] step=retry-goto-login url={LOGIN_URL} backoff={backoff_s}s")
            sleep(backoff_s)
            page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60_000)
            try:
                page.wait_for_load_state("networkidle", timeout=30_000)
            except Exception:  # noqa: BLE001
                pass
            print(f"[adp_login] step=retry-settled url={page.url}")

    if _is_throttled(page):
        raise AdpLoginThrottled(
            f"[adp_login] ADP throttle interstitial (sorry.adp.com) persisted across "
            f"all {total_attempts} login attempts — graceful ADP skip; next nightly retries. "
            f"URL: {page.url}"
        )
    raise RuntimeError(
        f"ADP login form did not render after {total_attempts} attempts "
        f"(User ID textbox never became visible). URL: {page.url}"
    )


def _ensure_logged_in(page, *, store: str, timeout_ms: int = 60_000) -> None:
    """Open ADP and complete a FRESH login using keychain creds.

    Stateless: assumes no cookies (ephemeral browser context). Flow:
        1. Navigate to LOGIN_URL.
        2. If we land directly on the dashboard (rare; only if the IP happens
           to have a residual ADP cookie), return.
        3. Fill User ID, click Next.
        4. Fill Password, click Sign in.
        5. Wait for the dashboard URL.

    Raises with detailed context on:
        - 2FA / verification challenges (Slack-DM operator before raising)
        - "User ID and/or password incorrect" — keychain creds need rotation
        - Any unexpected post-login URL
    """
    _load_login_page(page, timeout_ms=timeout_ms)

    if POST_LOGIN_URL_RE.search(page.url):
        return  # Already authenticated (shouldn't happen in ephemeral, but defensive).

    # Step 1: User ID. sdf-input wraps a real <input>; get_by_role finds it.
    uid_box = _wait_for_login_form(page)
    # The login page carries any scheduled-maintenance banner; capture the window
    # end NOW (it's gone once a post-login redirect bounces us to sorry.adp.com).
    maintenance_end = _read_maintenance_end(page)
    uid_box.fill(_get_adp_username(store))
    page.get_by_role("button", name=re.compile(r"^Next$", re.I)).first.click()
    print(f"[adp_login] step=clicked-next url={page.url}")

    # Step 2: Password. ADP enables the button only after the password field
    # is populated, so just press Enter rather than racing the disabled state.
    pw_box = page.get_by_role("textbox", name=re.compile(r"^Password$", re.I)).first
    try:
        pw_box.wait_for(state="visible", timeout=30_000)
    except Exception:
        # Could be MFA / device-trust challenge inserted between User ID and Password.
        _raise_with_evidence(
            page, store=store,
            reason="ADP did not show the Password field after User ID step. "
                   "Likely a 2FA / device-trust challenge that isn't wired yet. "
                   "Screenshot saved for diagnosis.",
        )
    pw_box.fill(_get_adp_password(store))
    page.keyboard.press("Enter")

    # Step 3: Land on dashboard. ADP may insert a verification step for 2FA,
    # or surface "User ID/password incorrect" inline.
    try:
        page.wait_for_url(POST_LOGIN_URL_RE, timeout=timeout_ms)
    except Exception:
        # Diagnose: incorrect creds, 2FA, or unexpected redirect.
        url = page.url.lower()
        on_2fa_url = any(
            t in url
            for t in ("verify", "challenge", "step-up", "mfa", "mobile_collection", "channel")
        )
        # Fallback: ADP sometimes serves 2FA on the bare login origin without a
        # URL marker — check the visible body for the "Verify your identity"
        # headline before deciding it's not 2FA.
        on_2fa_body = False
        if not on_2fa_url:
            try:
                on_2fa_body = page.get_by_text(
                    re.compile(r"verify your identity|verification code|text message", re.I)
                ).first.is_visible(timeout=1_000)
            except Exception:  # noqa: BLE001
                pass
        if on_2fa_url or on_2fa_body:
            print(f"[adp 2fa] detected challenge at {page.url}; engaging Slack-OTP handler...")
            _handle_adp_two_factor(page, store=store)
            # Handler returned: re-confirm we reached the dashboard.
            page.wait_for_url(POST_LOGIN_URL_RE, timeout=timeout_ms)
            return
        # Check for the inline "incorrect credentials" banner.
        try:
            banner = page.get_by_text(re.compile(r"user ID and/or password are incorrect", re.I)).first
            if banner.is_visible(timeout=2_000):
                _raise_with_evidence(
                    page, store=store,
                    reason=f"ADP rejected creds for {_get_adp_username(store)!r}. "
                           f"Keychain password is stale — rotate with: "
                           f"security add-generic-password -U -s jarvis-adp-{store} -a <user> -w <new-password>",
                )
        except Exception:  # noqa: BLE001
            pass
        # ADP serves a maintenance/throttle interstitial *after* a valid login —
        # sorry.adp.com (throttle) OR runpayroll.adp.com/public/maintenance/
        # maintenance.html (scheduled RUN maintenance) — distinct from the
        # login-form throttle handled in _wait_for_login_form. Treat it as a
        # graceful skip (AdpLoginThrottled → daily_refresh._handle_adp_throttle_skip)
        # rather than a hard failure, so the nightly exits 0 + alerts and a smart
        # retry auto-recovers. retry_at = parsed window-end + buffer when the banner
        # published one (login page OR this interstitial); else a fixed backoff.
        if _is_maintenance_interstitial(page.url):
            from agents.bhaga.scripts.otp_gate import AdpLoginThrottled  # local import
            from skills.adp_run_automation.maintenance import (
                compute_retry_at,
                default_retry_at,
            )
            end = maintenance_end or _read_maintenance_end(page)
            retry_at = compute_retry_at(end) if end else default_retry_at()
            basis = "window-end+buffer" if end else "no published end → fixed backoff"
            reason = (
                f"ADP served a maintenance/throttle interstitial after login "
                f"(url={page.url}); smart-retry scheduled for {retry_at.isoformat()} "
                f"({basis})"
            )
            print(f"[adp_login] step=post-login-maintenance url={page.url} "
                  f"retry_at={retry_at.isoformat()} basis={basis}")
            _raise_with_evidence(
                page, store=store, reason=reason,
                exc_factory=lambda msg: AdpLoginThrottled(msg, retry_at=retry_at),
            )
        _raise_with_evidence(
            page, store=store,
            reason=f"ADP login did not reach dashboard. Current URL: {page.url}",
        )


def _dismiss_adp_channel_collection(page) -> bool:
    """Skip ADP's "verify/add your mobile number" interstitial if it's showing.

    ADP RUN began inserting a contact-channel-collection page (custom
    ``<sdf-phone-number-input>`` + a "Verify mobile number" / "Remind me later"
    pair, ids ``channel-collection-save`` / ``channel-collection-remind``)
    between the password step and the actual OTP challenge. It carries no
    code-entry box, so the OTP handler would land here and raise "selector
    drift". We click "Remind me later" to dismiss it and let ADP continue.

    Returns True if the interstitial was detected and dismissed, else False.
    Never raises — a miss just means we weren't on that page.
    """
    # Detect by the page-unique elements (id is stable; fall back to role/text).
    detectors = (
        lambda: page.locator("#channel-collection-remind").first,
        lambda: page.locator("#phone-number-input").first,
        lambda: page.get_by_role("button", name=re.compile(r"remind me later", re.I)).first,
    )
    present = False
    for det in detectors:
        try:
            det().wait_for(state="visible", timeout=2_000)
            present = True
            break
        except Exception:  # noqa: BLE001
            continue
    if not present:
        return False

    print("[adp 2fa] mobile-number verify interstitial detected; clicking 'Remind me later'.")
    for clicker in (
        lambda: page.locator("#channel-collection-remind").first,
        lambda: page.get_by_role("button", name=re.compile(r"remind me later", re.I)).first,
    ):
        try:
            clicker().click(timeout=4_000)
            page.wait_for_timeout(2_000)  # let the dismissal navigate/settle
            return True
        except Exception:  # noqa: BLE001
            continue
    # Detected the page but couldn't click the skip control; report not-dismissed
    # so the caller's OTP flow raises its own evidence breadcrumb.
    print("[adp 2fa] WARN: interstitial detected but 'Remind me later' not clickable.")
    return False


def _handle_adp_two_factor(page, *, store: str) -> None:
    """Drive ADP RUN's SMS-OTP 2FA flow with operator-in-the-loop via Slack.

    Mirrors skills/square_tips/runner.py::_handle_square_two_factor. ADP's
    challenge page is built with custom <sdf-*> Web Components ("Secure
    Design Framework") and tends to drift in attribute structure, so this
    code prefers role/text matching over sdf-tag CSS selectors.

    Flow:
        1. Pick the "Text message" delivery option (radio or button).
        2. Click Continue/Send -> ADP sends SMS to the phone on file
           (per operator policy: number ending 0038 is first preference).
        3. Wait for the 6-digit code input on the next screen.
        4. Call skills.slack.adapter.request_otp(agent="bhaga") -> blocks
           up to 30 min waiting for the operator to reply with the code
           in the BHAGA DM (Socket-Mode listener picks it up).
        5. Fill the code, submit (Enter / Verify / Continue).

    Raises with screenshot+html evidence if any step's selector chain fails
    so we don't end up in a silent infinite wait.
    """
    print(f"[adp 2fa] starting handler; URL={page.url}")

    # Step 0: dismiss the "verify/add your mobile number" interstitial (ADP's
    # channel-collection page) if present. ADP began inserting this between
    # password submit and the OTP challenge; it has no code-entry box, so the
    # OTP flow below would mis-fire as "selector drift". Clicking "Remind me
    # later" skips it and lets ADP proceed to the real challenge / dashboard.
    if _dismiss_adp_channel_collection(page):
        # After skipping, we may already be on the dashboard (device trusted) —
        # if so there's no OTP to enter and we're done.
        try:
            page.wait_for_url(POST_LOGIN_URL_RE, timeout=15_000)
            print("[adp 2fa] reached dashboard after skipping mobile-verify; no OTP needed.")
            return
        except Exception:  # noqa: BLE001 — not on dashboard yet → continue OTP flow
            print("[adp 2fa] mobile-verify skipped; continuing to OTP challenge.")

    # Step 1: pick text-message delivery. ADP's picker varies between sdf-radio
    # and clickable sdf-card. Try several patterns.
    sms_picked = False
    for selector_fn in (
        lambda: page.get_by_role("radio", name=re.compile(r"text\s*message|SMS|text\b", re.I)).first,
        lambda: page.get_by_role("button", name=re.compile(r"text\s*message|SMS", re.I)).first,
        lambda: page.get_by_text(re.compile(r"text\s*message", re.I)).first,
    ):
        try:
            loc = selector_fn()
            loc.wait_for(state="visible", timeout=4_000)
            try:
                loc.check()
            except Exception:  # noqa: BLE001 — non-radio elements need click
                loc.click()
            sms_picked = True
            break
        except Exception:  # noqa: BLE001
            continue
    if not sms_picked:
        # ADP may have already moved straight to the code-entry screen (some
        # accounts skip the delivery picker). Don't fail — proceed and let
        # the code-input wait below confirm where we are.
        print("[adp 2fa] no delivery picker found; assuming code-entry screen already shown")

    # Step 2: click Continue / Send / Next to trigger SMS send. Skip if we
    # never found a picker — we're already on the code screen.
    if sms_picked:
        try:
            page.get_by_role(
                "button", name=re.compile(r"^continue$|^next$|^send$|^send code$", re.I)
            ).first.click(timeout=4_000)
        except Exception:  # noqa: BLE001
            try:
                page.keyboard.press("Enter")
            except Exception:  # noqa: BLE001
                pass

    # Step 3: wait for the code-entry input. ADP uses either a single text
    # input or a 6-digit-box widget. Cover both.
    page.wait_for_timeout(2_500)  # let the next screen render
    code_input = None
    for css in [
        "input[autocomplete='one-time-code']",
        "input[name='code']",
        "input[name='otp']",
        "input[inputmode='numeric']",
        "input[type='text'][maxlength='6']",
        "input[type='tel'][maxlength='6']",
    ]:
        try:
            loc = page.locator(css).first
            loc.wait_for(state="visible", timeout=4_000)
            code_input = loc
            break
        except Exception:  # noqa: BLE001
            continue

    six_digit_boxes = None
    if code_input is None:
        # Try the per-digit input widget.
        digit_inputs = page.locator("input[type='text'][maxlength='1'], input[type='tel'][maxlength='1']")
        try:
            count = digit_inputs.count()
        except Exception:  # noqa: BLE001
            count = 0
        if count >= 4:
            six_digit_boxes = digit_inputs
        else:
            _raise_with_evidence(
                page, store=store,
                reason=f"ADP 2FA code-entry input not found. Selector drift. "
                       f"Current URL: {page.url}",
            )

    # Step 4: request OTP via Slack DM, block until operator replies.
    from skills.slack.adapter import request_otp  # local import: optional dep
    from agents.bhaga.scripts.otp_gate import OtpWaitTimeout  # local import

    # In inline-autostart mode the gate already returned PROCEED and set
    # BHAGA_OTP_WAIT_S=900. Standalone/supervised callers with no env set keep
    # the generous 1800 s default.
    wait_s = int(os.environ.get("BHAGA_OTP_WAIT_S", "1800"))
    print(f"[adp 2fa] requesting OTP via Slack for store={store!r} (wait={wait_s}s); SMS expected at +1-XXX-XXX-0038")
    code = request_otp(
        user_id="U0APJRE5DC4",       # operator (primary_user_id from config.yaml)
        portal_name="ADP",
        timeout_seconds=wait_s,
        phone_hint="+1-XXX-XXX-0038",
        agent="bhaga",
    )
    if not code:
        raise OtpWaitTimeout(
            f"ADP 2FA: operator did not reply with the OTP within {wait_s}s. "
            "ADP step will be skipped; next nightly will retry with a fresh SMS."
        )
    code = code.strip().replace(" ", "").replace("-", "")
    print(f"[adp 2fa] got code (len={len(code)}); submitting.")

    # Step 5: fill the code and submit.
    if code_input is not None:
        code_input.fill(code)
    else:
        for i, ch in enumerate(code):
            six_digit_boxes.nth(i).fill(ch)

    try:
        page.get_by_role(
            "button", name=re.compile(r"^verify$|^continue$|^submit$|^next$|^sign in$", re.I)
        ).first.click(timeout=4_000)
    except Exception:  # noqa: BLE001
        page.keyboard.press("Enter")

    # Brief wait so the post-submit nav has a chance to start before the
    # caller's wait_for_url runs (avoids racing on stale `page.url`).
    page.wait_for_timeout(1_500)


def _target_meta_path(xlsx_path: pathlib.Path) -> pathlib.Path:
    """Sidecar path used to record which ``target_date`` an XLSX was scraped for.

    The XLSX itself is named by *download date* (``Timecard-2026-05-23.xlsx``),
    NOT by the business ``target_date`` the run was asking for. That means an
    XLSX downloaded earlier today for a DIFFERENT target_date can look
    indistinguishable from one downloaded just now for the current target_date.

    The 2026-05-23 silent-partial-success bug hit exactly that case: an orphan
    run for ``--date 2026-05-21`` wrote ``Timecard-2026-05-23.xlsx`` at 12:08,
    and the later run for ``--date 2026-05-22`` saw the file as "fresh on disk"
    and skipped the ADP scrape — possibly missing 5/22 punches.

    Fix: every successful download writes a small JSON sidecar
    (``Timecard-<today>.target-meta.json``) recording the target_date the file
    was scraped for. ``_xlsx_fresh_for_target`` requires BOTH the file to be
    fresh AND the sidecar's target_date to match the current run's target_date
    before declaring the file usable.
    """
    return xlsx_path.with_suffix(xlsx_path.suffix + ".target-meta.json")


def _write_target_meta(xlsx_path: pathlib.Path, target_date: Optional[datetime.date]) -> None:
    """Best-effort write of the target-date sidecar (see ``_target_meta_path``)."""
    try:
        meta = {
            "target_date": target_date.isoformat() if target_date else None,
            "downloaded_at": datetime.datetime.now().isoformat(),
            "xlsx_filename": xlsx_path.name,
        }
        _target_meta_path(xlsx_path).write_text(json.dumps(meta))
    except Exception as exc:  # noqa: BLE001
        print(f"[adp] WARN: could not write target-meta sidecar for {xlsx_path.name}: {exc}")


def _xlsx_fresh_for_target(
    xlsx_path: pathlib.Path,
    *,
    target_date: Optional[datetime.date],
    min_bytes: int,
) -> bool:
    """Tighter form of ``is_fresh_download`` that also checks the sidecar.

    Returns True iff:
      1. ``is_fresh_download(xlsx_path, min_bytes=...)`` says yes (file
         exists, mtime > CT-midnight today, size >= min_bytes), AND
      2. ``<xlsx_path>.target-meta.json`` exists and its ``target_date``
         field matches the current run's ``target_date``.

    Conservatively rejects when the sidecar is missing (older files from
    before this guard landed) so the next run re-scrapes rather than
    silently reusing a possibly-wrong-target XLSX. This means we eat at
    most ONE extra scrape after deploying this fix; subsequent runs sit
    fast.
    """
    if not is_fresh_download(xlsx_path, min_bytes=min_bytes):
        return False
    meta_path = _target_meta_path(xlsx_path)
    if not meta_path.exists():
        print(
            f"[adp] freshness: {xlsx_path.name} is on disk but has no target-meta "
            f"sidecar — treating as stale (need to record target_date for safe reuse)"
        )
        return False
    try:
        meta = json.loads(meta_path.read_text())
    except Exception as exc:  # noqa: BLE001
        print(f"[adp] freshness: target-meta sidecar unreadable ({exc}); treating as stale")
        return False
    want = target_date.isoformat() if target_date else None
    got = meta.get("target_date")
    if got != want:
        print(
            f"[adp] freshness: {xlsx_path.name} was scraped for target_date={got!r} "
            f"but this run wants target_date={want!r} — treating as stale"
        )
        return False
    return True


def _mark_run_step_done(
    step_name: str,
    *,
    refresh_date: Optional[datetime.date] = None,
    note: str = "",
) -> None:
    """Best-effort write of ~/.bhaga/state/run-<refresh_date>/{step_name}.done.

    Mirrors agents.bhaga.scripts.daily_refresh.mark_step_done so the bundle
    helper can record per-component completion (adp_timecard, adp_earnings)
    even though the orchestrator's run_step only sees the bundle-level call
    (adp_reports). Preserves per-component granularity for the wrapper
    roll-up alert and operator-facing debugging.

    The marker dir is keyed by ``refresh_date`` (the business date whose
    data we're publishing), NOT by today_ct. See daily_refresh._run_state_dir
    for the rationale. Falls back to today_ct only when refresh_date is
    None (e.g. legacy callers that haven't been threaded yet).
    """
    try:
        from zoneinfo import ZoneInfo
        ct = ZoneInfo("America/Chicago")
        key_date = refresh_date if refresh_date is not None else datetime.datetime.now(ct).date()
        d = pathlib.Path.home() / ".bhaga" / "state" / f"run-{key_date.isoformat()}"
        d.mkdir(parents=True, exist_ok=True)
        body = datetime.datetime.now(ct).isoformat()
        if note:
            body += f"\nnote: {note}"
        (d / f"{step_name}.done").write_text(body)
    except Exception as exc:  # noqa: BLE001
        print(f"[adp_bundle] WARN: could not write {step_name} marker: {exc}")


def _is_maintenance_interstitial(url: str) -> bool:
    """True if a post-login URL is an ADP maintenance / throttle interstitial.

    ADP serves at least two distinct pages for "service unavailable":
      - https://sorry.adp.com/sorry/                     (throttle / sorry)
      - https://runpayroll.adp.com/public/maintenance/maintenance.html  (RUN maintenance)
    Either means we did NOT reach the dashboard for a transient upstream reason —
    treat as a graceful skip (+ smart retry), never a hard scrape failure.
    """
    u = (url or "").lower()
    return ("sorry.adp.com" in u) or ("maintenance.html" in u) or ("/maintenance/" in u)


def _read_maintenance_end(page):
    """Best-effort: parse ADP's maintenance-window END from the page banner.

    Returns a UTC-aware datetime (window end), or None if no banner / unparseable.
    Never raises — a miss just means no smart retry is scheduled (next nightly
    re-attempts). Call this while the login page is visible (the banner shows
    there); the post-login sorry.adp.com page usually does not carry the times.
    """
    try:
        import datetime as _dt

        from skills.adp_run_automation.maintenance import parse_maintenance_end
        text = page.inner_text("body", timeout=2_000)
        return parse_maintenance_end(text, now=_dt.datetime.now(_dt.timezone.utc))
    except Exception:  # noqa: BLE001
        return None


def _raise_with_evidence(page, *, store: str, reason: str, exc_factory=RuntimeError) -> None:
    """Save screenshot + URL alongside the raise so failures are debuggable.

    The standard _browser_runtime evidence capture also fires when the
    exception propagates out of the launch_persistent with-block, so we
    end up with two snapshots — the one taken here is at the exact moment
    of the auth failure (most useful), the other is at context teardown.

    ``exc_factory`` lets the caller raise a typed exception (e.g.
    ``AdpLoginThrottled`` for a sorry.adp.com / maintenance redirect) while
    reusing the same screenshot-capture path; defaults to ``RuntimeError``.
    """
    try:
        ts = subprocess.run(["date", "+%Y%m%d-%H%M%S"], capture_output=True, text=True).stdout.strip()
        snap = f"/Users/adityaparikh/.bhaga/state/screenshots/adp-{store}-authfail-{ts}.png"
        page.screenshot(path=snap, full_page=True)
        reason += f"\nScreenshot: {snap}"
    except Exception:  # noqa: BLE001
        pass
    raise exc_factory(reason)


# ── Shared Reports navigation ─────────────────────────────────────


def _navigate_to_reports_landing(page, *, max_retries: int = 2) -> None:
    """Navigate from the ADP dashboard to the Reports listing page.

    Used by both _timecard_within_session and _earnings_within_session.

    Handles two failure modes observed in production (5/24-5/25):
      A) ADP lands on HOME after Reports-btn click — the "View all reports"
         link exists in a Reports widget card below the fold. wait_for(visible)
         times out because the element needs scrolling first.
      B) SPA navigation fails silently — main content area renders blank
         (sidebar OK, content white). The element never appears because JS
         hydration failed or an API call timed out.

    Strategy: scroll-first (attached → scroll → visible) + reload-retry on
    blank content.

    Post-condition: the "All Reports" listing is loaded (the page showing
    Time / Custom / Benefits / H-R / etc. accordion sections).
    """
    tag = "adp_reports"

    for attempt in range(1, max_retries + 2):
        print(f"[{tag}] step=click-reports-btn (attempt {attempt})")
        page.locator('[data-test-id="Reports-btn"]').first.click()
        page.wait_for_timeout(3_000)
        print(f"[{tag}] step=wait-content (attempt {attempt}) url={page.url}")

        content_indicators = page.locator(
            '[data-test-id="view-all-reports"], '
            '[data-test-id="reports-tile-view-all-reports-button"], '
            '[data-test-id*="section_Head"]'
        )
        try:
            content_count = content_indicators.count()
        except Exception:  # noqa: BLE001
            content_count = 0

        if content_count > 0:
            print(f"[{tag}] step=content-loaded (attempt {attempt}, "
                  f"indicators={content_count})")
            break

        if attempt <= max_retries:
            print(f"[{tag}] step=content-blank, reloading page (attempt {attempt})")
            page.reload(wait_until="domcontentloaded", timeout=30_000)
            try:
                page.wait_for_load_state("networkidle", timeout=8_000)
            except Exception:  # noqa: BLE001
                pass
            page.wait_for_timeout(2_000)
        else:
            print(f"[{tag}] step=content-blank after {max_retries} retries; "
                  f"proceeding anyway")

    view_all = page.locator(
        '[data-test-id="view-all-reports"], '
        '[data-test-id="reports-tile-view-all-reports-button"]'
    ).first

    print(f"[{tag}] step=view-all-reports wait-attached")
    view_all.wait_for(state="attached", timeout=15_000)

    print(f"[{tag}] step=view-all-reports scroll-into-view")
    view_all.scroll_into_view_if_needed(timeout=5_000)
    view_all.wait_for(state="visible", timeout=5_000)

    print(f"[{tag}] step=view-all-reports click")
    view_all.click()
    page.wait_for_timeout(2_000)
    print(f"[{tag}] step=view-all-reports clicked url={page.url}")


# ── Timecard scrape ────────────────────────────────────────────────


def download_timecard(
    *,
    store: str = "palmetto",
    target_date: Optional[datetime.date] = None,
    headed: bool = True,
    slow_mo_ms: int = 50,
    keep_open_on_error: bool = False,
) -> pathlib.Path:
    """Open Reports > Time reports > Timecard, select pay periods, apply
    changes, click Export to Excel, save .xlsx.

    Args:
        target_date: if provided, only the single pay period whose date range
            contains this date is selected (nightly-incremental mode). If
            None, ALL pay periods are selected (backfill mode — backwards
            compatible with the original behavior).

    Idempotency: if today's Timecard XLSX is already on disk (CT-today mtime),
    skip the browser entirely and return the cached path. Eliminates
    duplicate ADP 2FA SMS on cron retries.
    """
    expected = DOWNLOADS_DIR / f"Timecard-{datetime.date.today().isoformat()}.xlsx"
    if _xlsx_fresh_for_target(expected, target_date=target_date, min_bytes=10_000):
        print(f"[adp_timecard] SKIP browser — fresh Timecard XLSX already on disk: {expected}")
        return expected

    with launch_persistent(
        portal="adp",
        headed=headed,
        slow_mo_ms=slow_mo_ms,
        keep_open_on_error=keep_open_on_error,
    ) as (ctx, page):
        _ensure_logged_in(page, store=store)
        path = _timecard_within_session(page, target_date=target_date, store=store)
        _write_target_meta(path, target_date)
        return path


def _load_store_profile(store: str) -> dict:
    return json.loads((STORE_PROFILES / f"{store}.json").read_text())


def _timecard_within_session(
    page,
    *,
    target_date: Optional[datetime.date] = None,
    store: str = "palmetto",
) -> pathlib.Path:
    """Run the Timecard scrape on an already-authenticated ADP dashboard page.

    Extracted from download_timecard so the bundle helper can drive both
    Timecard and Earnings within a single browser session (one login, one
    OTP cost). Caller owns the browser context lifecycle.

    Pre-condition: `page` is on the v2 ADP RUN dashboard (POST_LOGIN_URL_RE).
    """
    _navigate_to_reports_landing(page)
    # Single Reports modal: ensure the "Time" section header is loaded, then
    # click the Timecard report tile. The section accordion is often already
    # expanded by default; if not, clicking the header expands it.
    time_section_header = page.locator('[data-test-id="Time-section_Head_HeaderLabel"]').first
    time_section_header.wait_for(state="visible", timeout=15_000)
    timecard_tile = page.locator('[data-test-id="Time-tile-list-item-Timecard"]').first
    try:
        timecard_tile.wait_for(state="visible", timeout=3_000)
    except Exception:
        time_section_header.click()
        page.wait_for_timeout(500)
        timecard_tile.wait_for(state="visible", timeout=10_000)
    timecard_tile.scroll_into_view_if_needed(timeout=5_000)
    timecard_tile.click()
    # Timecard report iframe rendering — wait for the iframe directly.
    page.locator("iframe[name='mdfTimeFrame']").wait_for(state="attached", timeout=20_000)
    page.wait_for_timeout(2_500)

    # All Timecard report controls live inside iframe[name="mdfTimeFrame"].
    page.wait_for_timeout(4_000)
    frame = page.frame_locator("iframe[name='mdfTimeFrame']")

    # Pay Period selection. There are TWO comboboxes with "Pay Period" in
    # their name — "Report Period" (single-select; its current value is
    # "Pay P..." which a permissive regex matches) and the actual "Pay Period"
    # multi-select. Anchor on exact name to disambiguate.
    # Accessible name format on these sdf-inputs is "{Label} {Value}", e.g.
    # "Report Period Pay P..." or "Pay Period 1 Selected". Anchor START so we
    # don't accidentally match Report Period (whose value contains "Pay P").
    pp_combobox = frame.get_by_role(
        "combobox", name=re.compile(r"^Pay Period\b", re.I)
    ).first
    pp_combobox.click()
    page.wait_for_timeout(800)

    if target_date is not None:
        # Nightly-incremental: one period covering target_date (not Select All).
        _select_pay_period_for_target(
            frame, page, target_date=target_date, store=store
        )
    else:
        # Backfill / default mode: select every pay period exposed in the
        # dropdown. The multi-select listbox has a "Select All" checkbox at
        # the top; fall back to per-option clicks if not exposed.
        try:
            select_all = frame.get_by_role(
                "option", name=re.compile(r"^Select All$", re.I)
            ).first
            select_all.wait_for(state="visible", timeout=5_000)
            select_all.click()
        except Exception:
            opts = frame.get_by_role(
                "option", name=_PAY_PERIOD_DATE_RANGE_RE
            )
            n = opts.count()
            for i in range(n):
                opts.nth(i).click()
    page.wait_for_timeout(500)
    page.keyboard.press("Escape")
    page.wait_for_timeout(300)

    # Format: must be "Continuous" — "Page Layout" only exposes Print (PDF)
    # and hides the Excel export button entirely.
    fmt_combobox = frame.get_by_role(
        "combobox", name=re.compile(r"^Format\b", re.I)
    ).first
    fmt_combobox.click()
    page.wait_for_timeout(500)
    try:
        frame.get_by_role(
            "option", name=re.compile(r"^Continuous$", re.I)
        ).first.click()
    except Exception:
        page.keyboard.press("ArrowDown")
        page.keyboard.press("Enter")
    page.wait_for_timeout(400)

    frame.get_by_role("button", name=re.compile(r"^Apply Changes$", re.I)).first.click()
    page.wait_for_timeout(8_000)

    # Export To Excel — the SDF-BUTTON custom element.
    excel_btn_selector = "#report-excel-button"
    try:
        frame.locator(excel_btn_selector).first.wait_for(state="visible", timeout=10_000)
    except Exception:
        page.wait_for_timeout(5_000)
        frame.locator(excel_btn_selector).first.wait_for(state="visible", timeout=10_000)

    rename = f"Timecard-{datetime.date.today().isoformat()}.xlsx"
    path = download_to(
        page,
        trigger=lambda: frame.locator(excel_btn_selector).first.click(),
        rename_to=rename,
        timeout_ms=_TIMECARD_DOWNLOAD_TIMEOUT_MS,
    )
    return path


# ── Earnings scrape ────────────────────────────────────────────────


def download_earnings(
    *,
    store: str = "palmetto",
    start_date: Optional[datetime.date] = None,
    end_date: Optional[datetime.date] = None,
    headed: bool = True,
    slow_mo_ms: int = 50,
    keep_open_on_error: bool = False,
) -> pathlib.Path:
    """Open Reports > My saved custom reports > '{wage_rate_report_name}',
    set Custom date range, preview, download Excel.

    Default date range: last 90 days (more than enough to infer current rates
    AND capture the most recent pay-period's Credit Card Tips Owed).

    Idempotency: if today's Earnings-and-Hours XLSX is already on disk
    (CT-today mtime), skip the browser entirely and return the cached path.
    """
    today = datetime.date.today()
    expected = DOWNLOADS_DIR / f"Earnings-and-Hours-V1-{today.isoformat()}.xlsx"
    # Standalone download_earnings has no target_date concept — the report
    # is preset-driven ("Last payroll"). Use end_date if provided, else
    # today_ct, as the freshness key so a same-day rerun with a different
    # window doesn't silently reuse the prior XLSX.
    target_for_freshness = end_date or today
    if _xlsx_fresh_for_target(expected, target_date=target_for_freshness, min_bytes=5_000):
        print(f"[adp_earnings] SKIP browser — fresh Earnings XLSX already on disk: {expected}")
        return expected

    profile = _load_store_profile(store)
    report_name = profile["adp_run"].get("wage_rate_report_name", "Earnings and Hours V1")
    start = start_date or (today - datetime.timedelta(days=90))
    end = end_date or today

    with launch_persistent(
        portal="adp",
        headed=headed,
        slow_mo_ms=slow_mo_ms,
        keep_open_on_error=keep_open_on_error,
    ) as (ctx, page):
        _ensure_logged_in(page, store=store)
        path = _earnings_within_session(
            page, store=store, start=start, end=end
        )
        _write_target_meta(path, target_for_freshness)
        return path


def _assert_earnings_xlsx_has_rows(path: pathlib.Path) -> int:
    """Assert that the downloaded Earnings & Hours XLSX contains data rows.

    The ADP report layout (verified 2026-05-19) has a 5-line preamble
    followed by a header row, e.g.:
        row 0: "Earnings and Hours V1"
        row 1: "Company: <name>"
        row 2: "IID: <num>"
        row 3: "DateRange : <start> to <end>"
        row 4: "Report Generated On: <ts>"
        row 5: "Employee Name" "Payroll Check Date" ... (header row)
        row 6+: data rows
    We count anything with a non-null first-column value AFTER the header
    row, plus any row whose 7th column ("Payroll Earning Amount") has a
    numeric value — that catches the continuation rows where
    Employee Name is blank (same employee, additional earning lines like
    "Credit Card Tips Owed" or "Cash tips") but the row is still a
    legitimate earnings entry we must keep.

    Returns the data-row count. Raises RuntimeError if 0.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    data_rows = 0
    seen_header = False
    for row in ws.iter_rows(values_only=True):
        if not seen_header:
            if row and row[0] == "Employee Name":
                seen_header = True
            continue
        # After header: a row counts as data if column[0] (employee name)
        # is set OR if column[6] (Payroll Earning Amount) is a number.
        # This handles ADP's "blank-employee-name continuation row"
        # pattern for multi-line earnings within a single employee+period.
        first = row[0] if len(row) > 0 else None
        amount = row[6] if len(row) > 6 else None
        if first or isinstance(amount, (int, float)):
            data_rows += 1
    wb.close()
    if data_rows == 0:
        raise RuntimeError(
            f"Earnings & Hours export was empty (0 data rows) — "
            f"pay-period selector may have failed (open / in-flight "
            f"payroll has no closed-payroll data). XLSX: {path}"
        )
    return data_rows


def _wait_for_earnings_ready_button(page, *, timeout_ms: int, locator_specs: list) -> object:
    """Wait for the "Your report is ready to download" button to appear.

    ADP queues async report generation after the exportExcel click; the modal
    can take 3-90+ seconds depending on server load. This function polls a
    ranked list of locators until one becomes visible within *timeout_ms*, then
    returns that locator (bound to the matched element) for the caller to click.

    On total timeout, captures a full-page screenshot + HTML source into the
    standard diagnostic snapshot directory (mirrors the tile-missing block) and
    raises RuntimeError with the URL + selectors tried so future incidents are
    debuggable from GCS evidence without a live browser session.

    Locator specs: list of (selector_or_label, log_tag) pairs. The first entry
    uses `page.locator(selector)`, subsequent entries are resolved by their
    tag to support role-based locators.
    """
    import time as _time

    deadline = _time.monotonic() + timeout_ms / 1000.0
    poll_interval_s = 1.0

    def _resolve(spec):
        selector, tag = spec
        if tag == "role-button-download-report":
            return page.get_by_role("button", name=re.compile(r"Download report", re.I)).first
        return page.locator(selector).first

    tried = [tag for _, tag in locator_specs]
    matched_btn = None
    while _time.monotonic() < deadline:
        for spec in locator_specs:
            btn = _resolve(spec)
            try:
                if btn.is_visible():
                    matched_btn = btn
                    print(f"[earnings] step=ready-dialog-found selector={spec[1]!r}")
                    return matched_btn
            except Exception:  # noqa: BLE001
                pass
        _time.sleep(poll_interval_s)

    # Total timeout — save diagnostic snapshot then raise.
    try:
        import datetime as _dt
        ts = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        shot_dir = pathlib.Path.home() / ".bhaga" / "state" / "screenshots"
        shot_dir.mkdir(parents=True, exist_ok=True)
        page.screenshot(
            path=str(shot_dir / f"adp-earnings-ready-dialog-missing-{ts}.png"),
            full_page=True,
        )
        (shot_dir / f"adp-earnings-ready-dialog-missing-{ts}.html").write_text(page.content())
        print(f"[earnings] saved diagnostic snapshot: "
              f"adp-earnings-ready-dialog-missing-{ts}.{{png,html}}")
    except Exception as snap_exc:  # noqa: BLE001
        print(f"[earnings] could not save diagnostic snapshot: {snap_exc}")

    raise RuntimeError(
        f"[earnings] ready-dialog button never appeared after {timeout_ms}ms "
        f"(URL={page.url}). Selectors tried: {tried}. "
        f"This is likely ADP async report generation taking longer than expected "
        f"or a selector change on the 'Your report is ready to download' modal. "
        f"Check the diagnostic snapshot in ~/.bhaga/state/screenshots/ for the DOM state."
    )


def _earnings_within_session(
    page,
    *,
    store: str,
    target_date: Optional[datetime.date] = None,
    start: Optional[datetime.date] = None,
    end: Optional[datetime.date] = None,
    use_custom_range: bool = False,
) -> pathlib.Path:
    """Run the Earnings & Hours scrape on an already-authenticated page.

    Pre-condition: `page` is on the v2 ADP RUN dashboard. Caller owns the
    browser context.

    DECISION LOCK-IN: two modes
    - use_custom_range=False (nightly incremental): selects "Last payroll" preset
      — the only preset guaranteed non-empty for the current pay cycle.
    - use_custom_range=True (explicit historical window): selects "Custom date range",
      fills From/To (MM/DD/YYYY), and sets Employment status = All so terminated
      employees appear in the historical window. Raises if start/end missing or
      range > 12 months (ADP server cap).

    Flow (verified end-to-end 2026-05-19):

        1.  Click Reports-btn (top nav).
        2.  Click view-all-reports (or its homepage-widget twin).
        3.  Expand the "Custom" accordion section — tiles are CSS-hidden
            until the header is clicked, despite being in the DOM.
        4.  Click the saved "Earnings and Hours V1" tile (per-store
            numeric suffix; anchor on the label).
        5.  Open date-range-field dropdown:
            - nightly: select "Last payroll"
            - historical: select "Custom date range", fill From/To,
              set Employment status = All
        6.  Click view-custom-report (Preview).
        7.  Pre-flight: check the AG-Grid for "No Rows To Show".
        8.  Click Download → Excel (.xlsx) → Download report.
        9.  Wait for the file to land via `download_to`.
        10. Post-flight: assert ≥1 data row in XLSX.
    """
    profile = _load_store_profile(store)
    report_name = profile["adp_run"].get("wage_rate_report_name", "Earnings and Hours V1")

    print(f"[earnings] step=pre-reports url={page.url} "
          f"target_date={target_date.isoformat() if target_date else 'none'} "
          f"use_custom_range={use_custom_range} start={start} end={end}")
    _navigate_to_reports_landing(page)
    print(f"[earnings] step=after-view-all-reports url={page.url}")

    # The Reports landing has accordion sections (verified 2026-05-19):
    #   Time / Custom / Benefits / H-R / Misc / Payroll / Taxes
    # Saved-reports tiles ([data-test-id="Custom-tile-list-item-<id>"]) live
    # inside the Custom section and start collapsed/hidden — Playwright
    # resolves them in the DOM but reports `hidden` until the header is
    # clicked. Mirrors download_timecard's expand-if-needed pattern.
    custom_header = page.locator(
        '[data-test-id="Custom-section_Head_HeaderLabel"]'
    ).first
    custom_header.wait_for(state="visible", timeout=15_000)

    # Tile selector: parent card filtered by an exact-text label child.
    # The label's data-test-id has a per-store numeric suffix (6358 on
    # palmetto) so anchor on the prefix and the label text — never the
    # literal id.
    report_tile = page.locator(
        '[data-test-id^="Custom-tile-list-item-"]'
    ).filter(
        has=page.locator('[data-test-id^="custom-report-label-"]').filter(
            has_text=re.compile(rf"^{re.escape(report_name)}$", re.I)
        )
    ).first
    try:
        report_tile.wait_for(state="visible", timeout=3_000)
        print("[earnings] step=custom-section already-expanded; "
              "tile visible without click")
    except Exception:  # noqa: BLE001
        print("[earnings] step=expand-custom-section "
              "(tile not visible until accordion expanded)")
        custom_header.scroll_into_view_if_needed(timeout=3_000)
        custom_header.click()
        page.wait_for_timeout(800)

    try:
        report_tile.wait_for(state="visible", timeout=15_000)
    except Exception as exc:  # noqa: BLE001
        print(f"[earnings] tile STILL not visible after expanding Custom "
              f"section ({type(exc).__name__}); saving diagnostic snapshot")
        try:
            ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            shot_dir = pathlib.Path.home() / ".bhaga" / "state" / "screenshots"
            shot_dir.mkdir(parents=True, exist_ok=True)
            page.screenshot(
                path=str(shot_dir / f"adp-earnings-tile-missing-{ts}.png"),
                full_page=True,
            )
            (shot_dir / f"adp-earnings-tile-missing-{ts}.html").write_text(page.content())
            print(f"[earnings] saved diagnostic snapshot: "
                  f"adp-earnings-tile-missing-{ts}.{{png,html}}")
        except Exception as exc3:  # noqa: BLE001
            print(f"[earnings] could not save diagnostic snapshot: {exc3}")
        raise exc

    report_tile.scroll_into_view_if_needed(timeout=5_000)
    print(f"[earnings] step=click-saved-report name={report_name!r}")
    report_tile.click()

    # Custom report builder modal opens (NO iframe — main frame
    # `<dialog data-test-id="modal-dialog">`). Wait for the modal-level
    # date-range select to be ready, then select "Last payroll".
    date_range_select = page.locator('[data-test-id="date-range-field"]').first
    date_range_select.wait_for(state="visible", timeout=20_000)
    date_range_select.click()
    page.wait_for_timeout(500)
    print("[earnings] step=open-date-range-dropdown")

    # Listbox options exposed: Last month / Last year / Last quarter /
    # Custom date range / Last payroll.
    # nightly (use_custom_range=False): "Last payroll" — the most-recently-CLOSED
    #   payroll; guaranteed non-empty for current wage-rate + CC-Tips-Owed lines.
    # historical backfill (use_custom_range=True): "Custom date range" with explicit
    #   From/To + Employment status = All (includes terminated employees).
    if use_custom_range:
        if start is None or end is None:
            raise ValueError(
                "use_custom_range=True requires both start and end dates"
            )
        if (end - start).days > 366:
            raise ValueError(
                f"ADP earnings range >12 months not supported "
                f"({start}..{end}); chunk the backfill into ≤12-month windows."
            )
        page.get_by_role("option", name=re.compile(r"^Custom date range$", re.I)).first.click()
        page.wait_for_timeout(500)
        page.get_by_role("textbox", name=re.compile(r"^From", re.I)).first.fill(start.strftime("%m/%d/%Y"))
        page.get_by_role("textbox", name=re.compile(r"^To", re.I)).first.fill(end.strftime("%m/%d/%Y"))
        # Employment status = All — includes terminated employees in historical window.
        # Best-guess role-based locator; verified / corrected in T6 sandbox DOM snapshot.
        try:
            emp = page.get_by_role("combobox", name=re.compile(r"Employment status", re.I)).first
            emp.click()
            page.wait_for_timeout(300)
            page.get_by_role("option", name=re.compile(r"^All$", re.I)).first.click()
        except Exception as exc:  # noqa: BLE001
            print(
                f"[earnings] WARN: Employment-status=All not set ({exc}); "
                f"continuing with default. Verify selector via sandbox DOM snapshot."
            )
        print(f"[earnings] step=selected-custom-range {start}..{end} employment=All")
    else:
        page.get_by_role(
            "option", name=re.compile(r"^Last payroll$", re.I)
        ).first.click()
        print("[earnings] step=selected-last-payroll")
    page.wait_for_timeout(800)

    # Preview report — populates the AG-Grid. Wait ≥6s before checking
    # for empty-grid sentinel (network round-trip can be slow).
    page.locator('[data-test-id="view-custom-report"]').first.click()
    print("[earnings] step=clicked-preview-report")
    page.wait_for_timeout(7_000)

    # Pre-flight empty-grid check — raise BEFORE we trigger the heavier
    # async report generation. If "Last payroll" itself returned empty
    # (e.g. a brand-new store with no closed payrolls yet), surface that
    # immediately rather than letting the file land empty.
    grid_locator = page.locator('[data-test-id="custom-reporting-ag-grid"]').first
    try:
        grid_locator.wait_for(state="visible", timeout=10_000)
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(
            f"Earnings preview grid never rendered ({type(exc).__name__}). "
            f"URL={page.url}"
        )
    try:
        grid_text = grid_locator.inner_text(timeout=4_000)
    except Exception:  # noqa: BLE001
        grid_text = ""
    if "No Rows To Show" in grid_text:
        raise RuntimeError(
            "Earnings preview grid shows 'No Rows To Show' even with "
            "'Last payroll' filter — the most recent payroll appears empty "
            f"for store={store!r}. Aborting before download to avoid "
            f"shipping a zero-row XLSX."
        )

    # Download → Excel (.xlsx) submenu → confirmation modal.
    page.get_by_role("button", name=re.compile(r"^Download$", re.I)).first.click()
    page.wait_for_timeout(500)
    page.locator('[data-test-id="exportExcel"]').first.click()
    print("[earnings] step=clicked-export-excel")

    # "Your report is ready to download" focus-pane — ADP queues async report
    # generation after exportExcel; the modal can take up to ~90s on loaded
    # servers. We poll a ranked list of locators for the full timeout window so
    # a single slow generation cycle or minor selector rename doesn't abort the
    # nightly. Configurable via BHAGA_ADP_EARNINGS_READY_TIMEOUT_MS (default
    # 90 000 ms — raised from the original 45 000 ms that failed 2026-06-23).
    _ready_timeout_ms = int(
        os.environ.get("BHAGA_ADP_EARNINGS_READY_TIMEOUT_MS", "90000")
    )
    # Ranked fallback locators for the "Download report" button inside the
    # ready dialog. Try primary first; if the selector ever drifts, role-based
    # and dialog-scoped variants can still resolve.
    _ready_locator_specs = [
        ('[data-test-id="download-report"]', "data-test-id=download-report"),
        ('get_by_role("button", name=re.compile(r"Download report", re.I))', "role-button-download-report"),
        ('[aria-label="Download report"]', "aria-label=Download-report"),
    ]
    ready_dialog_btn = _wait_for_earnings_ready_button(
        page, timeout_ms=_ready_timeout_ms, locator_specs=_ready_locator_specs
    )
    print("[earnings] step=download-report-button-ready")

    rename = f"Earnings-and-Hours-V1-{datetime.date.today().isoformat()}.xlsx"
    path = download_to(
        page,
        trigger=lambda: ready_dialog_btn.click(),
        rename_to=rename,
        timeout_ms=90_000,
    )
    print(f"[earnings] step=downloaded path={path}")

    # Hard row-count guard — fail loudly rather than ship an empty file.
    # The pre-flight grid check above usually catches empties, but we
    # double-check the actual XLSX (defense in depth: ADP has been known
    # to serve a header-only XLSX even when the preview grid rendered).
    data_rows = _assert_earnings_xlsx_has_rows(path)
    print(f"[earnings] step=row-count-guard OK rows={data_rows} file={path}")

    return path


# ── Team Schedule scrape (DOM, no file export) ─────────────────────


def _open_team_schedule(page, *, timeout_ms: int = 30_000):
    """From the v2 dashboard, open "Manage Schedules" and return the grid frame.

    The home-page "Team Schedule" quick-action is a hidden template anchor
    ``<a id="TEMPUS_WEEKLY_SCHEDULE" href="#xfm-...">``; a normal click times
    out (not actionable), so we fire its handler via JS. The grid renders in
    ``iframe[name="timePartnerFrame"]``. Idempotent: if the grid frame is
    already present (schedule already open), we just return it.
    """
    from skills.adp_run_automation import schedule_backend as sb

    def _grid_frame():
        for fr in page.frames:
            if fr.name == sb.SCHEDULE_GRID_FRAME_NAME:
                return fr
        return None

    if _grid_frame() is None:
        print(f"[adp_schedule] step=open-team-schedule (#{sb.TEAM_SCHEDULE_ANCHOR_ID})")
        # TEMPUS anchors hydrate after Time / home widgets; nudge Time first.
        page.evaluate(
            """() => {
              const t = document.querySelector('[data-test-id="Time-btn"]')
                || [...document.querySelectorAll('button,a')].find(
                     e => (e.getAttribute('aria-label')||'') === 'Time'
                        || /^(Time)$/i.test((e.innerText||'').trim()));
              if (t) t.click();
            }"""
        )
        # TEMPUS anchors hydrate a beat after Time / home widgets.
        deadline_ids = time.monotonic() + 20.0
        while time.monotonic() < deadline_ids:
            page.wait_for_timeout(500)
            if page.evaluate(
                "(id) => !!document.getElementById(id)", sb.TEAM_SCHEDULE_ANCHOR_ID
            ):
                break
        clicked = page.evaluate(
            "(id) => { const a=document.getElementById(id); if(a){a.click(); return true;} return false; }",
            sb.TEAM_SCHEDULE_ANCHOR_ID,
        )
        if not clicked:
            _raise_with_evidence(
                page, store="palmetto",
                reason=f"Team Schedule anchor #{sb.TEAM_SCHEDULE_ANCHOR_ID} not found on dashboard.",
            )

    # Wait for the grid frame to attach and a footer total to render.
    deadline = time.monotonic() + timeout_ms / 1000.0
    frame = None
    while time.monotonic() < deadline:
        frame = _grid_frame()
        if frame is not None:
            try:
                if frame.evaluate(
                    "() => document.querySelectorAll('team-schedule-total').length > 0"
                ):
                    break
            except Exception:  # noqa: BLE001 — frame still navigating
                pass
        page.wait_for_timeout(500)
    if frame is None:
        _raise_with_evidence(
            page, store="palmetto",
            reason="Manage Schedules grid frame (timePartnerFrame) never appeared.",
        )
    print(f"[adp_schedule] step=grid-frame-ready name={frame.name}")
    return frame


def _scrape_one_week(page, frame) -> dict:
    """Read the current week's label + footer totals + per-employee day cells."""
    import re as _re

    from skills.adp_run_automation import schedule_backend as sb

    week_label = (
        frame.get_by_text(_re.compile(_re.escape(sb.WEEK_LABEL_TEXT)))
        .first.inner_text(timeout=10_000)
    )
    ext = frame.evaluate(sb.SCHEDULE_EXTRACT_JS)
    emp = {"employees": [], "headers": []}
    try:
        # Wait for employee rows to paint (footer totals appear first).
        deadline = time.monotonic() + 15.0
        while time.monotonic() < deadline:
            n = frame.evaluate(
                "() => document.querySelectorAll('.worker-name').length"
            )
            if n and n > 0:
                break
            page.wait_for_timeout(400)
        # Virtualized mid-list rows hydrate day cells only when scrolled into
        # view. Extract one calendar-row at a time, then reset scroll so the
        # next-week chevron click geometry stays stable.
        n_rows = frame.evaluate(
            "() => document.querySelectorAll('.calendar-row').length"
        ) or 0
        frame.evaluate("() => { window.__adpEmpExtract = { headers: null, employees: [] }; }")
        for i in range(int(n_rows)):
            frame.evaluate(
                """(i) => {
                  const rows = document.querySelectorAll('.calendar-row');
                  if (rows[i]) rows[i].scrollIntoView({ block: 'center' });
                }""",
                i,
            )
            page.wait_for_timeout(150)
            frame.evaluate(sb.SCHEDULE_EMPLOYEE_EXTRACT_ONE_JS, i)
        emp = frame.evaluate(
            """() => {
              const e = window.__adpEmpExtract || { headers: [], employees: [] };
              const headers = (e.headers || []).map(h =>
                (typeof h === 'string' ? h : (h && h.text)) || null
              ).filter(Boolean);
              return { headers, employees: e.employees || [] };
            }"""
        )
        frame.evaluate(
            """() => {
              const scroller = document.querySelector('.calendar-view')
                || document.querySelector('.team-work-schedules-list');
              if (scroller) scroller.scrollTop = 0;
              const h = document.querySelector('.header-row')
                || document.querySelector('.day-cell.column-header');
              if (h) h.scrollIntoView({ block: 'start' });
            }"""
        )
        page.wait_for_timeout(300)
    except Exception as exc:  # noqa: BLE001 — employee grain is additive
        print(f"[adp_schedule] WARN: employee extract failed: {exc!r}")
    print(
        f"[adp_schedule] employee_rows={len(emp.get('employees') or [])} "
        f"headers={emp.get('headers')}"
    )
    return {
        "week_label": week_label.strip(),
        "days": ext.get("days") or [],
        "grand": ext.get("grand"),
        "employee_rows": emp.get("employees") or [],
        "day_headers": emp.get("headers") or [],
    }


def _goto_next_week(page, frame) -> None:
    """Advance the schedule grid to the next week.

    The ‹ › chevrons live in Shadow DOM next to the week-label link. We locate
    the (shadow-piercing) week label and click just to the right of it, where
    the "next week" chevron sits.

    Two-phase wait to avoid a render race: the shadow-DOM week label updates
    almost immediately on click, but the <team-schedule-total> footer cells
    (light DOM) re-render a beat later. If we extract as soon as the label
    flips we capture the PREVIOUS week's totals. So we wait for (a) the label
    to change AND (b) the footer totals to change (or a settle timeout, which
    covers the rare case of two weeks with identical totals).
    """
    import re as _re

    from skills.adp_run_automation import schedule_backend as sb

    label = frame.get_by_text(_re.compile(_re.escape(sb.WEEK_LABEL_TEXT))).first
    before_label = label.inner_text(timeout=5_000).strip()
    before_totals = frame.evaluate(sb.SCHEDULE_EXTRACT_JS)
    box = label.bounding_box()
    if not box:
        raise RuntimeError("Could not locate the week-selector label to navigate weeks.")
    # Click the › chevron just to the right of the week label. Prefer a
    # locator-relative click (stable after employee-row hydrate scroll);
    # fall back to page mouse coords.
    try:
        label.click(
            position={"x": box["width"] + 18, "y": max(box["height"] / 2, 8)},
            force=True,
            timeout=5_000,
        )
    except Exception:  # noqa: BLE001
        page.mouse.click(box["x"] + box["width"] + 16, box["y"] + box["height"] / 2)

    # Phase 1: label must change (confirms the nav fired).
    deadline = time.monotonic() + 12.0
    label_changed = False
    while time.monotonic() < deadline:
        page.wait_for_timeout(300)
        try:
            now = label.inner_text(timeout=2_000).strip()
        except Exception:  # noqa: BLE001
            now = before_label
        if now != before_label:
            label_changed = True
            print(f"[adp_schedule] step=advanced-week {before_label!r} -> {now!r}")
            break
    if not label_changed:
        raise RuntimeError(
            f"Next-week navigation did not change the week label (still {before_label!r}). "
            "Chevron position may have drifted."
        )

    # Phase 2: footer totals must re-render. Poll until they differ from the
    # pre-nav snapshot; if they never differ within the settle window the two
    # weeks genuinely have identical totals — fine, proceed.
    settle_deadline = time.monotonic() + 6.0
    while time.monotonic() < settle_deadline:
        page.wait_for_timeout(300)
        try:
            now_totals = frame.evaluate(sb.SCHEDULE_EXTRACT_JS)
        except Exception:  # noqa: BLE001
            continue
        if now_totals != before_totals:
            return
    print("[adp_schedule] step=totals-settle-timeout (assuming identical-week totals)")


def _schedule_within_session(page, *, weeks: int = None) -> list[dict]:
    """Scrape `weeks` consecutive weeks of Team Schedule totals.

    Pre-condition: `page` is on the v2 ADP RUN dashboard (POST_LOGIN_URL_RE).
    Returns a list of per-week raw payloads (see schedule_backend.build_schedule_records).
    """
    from skills.adp_run_automation import schedule_backend as sb

    weeks = weeks or sb.DEFAULT_WEEKS
    frame = _open_team_schedule(page)
    payloads: list[dict] = []
    for i in range(weeks):
        payloads.append(_scrape_one_week(page, frame))
        if i < weeks - 1:
            _goto_next_week(page, frame)
    return payloads


def _write_schedule_json(payloads: list[dict], *, store: str) -> pathlib.Path:
    """Persist the scraped week payloads as Schedule-<today>.json in DOWNLOADS_DIR.

    Mirrors the timecard/earnings "drop a file in downloads/, parse it later in
    backfill_from_downloads" contract — keeps the scrape and the BQ load
    decoupled and the load step unit-testable off a fixture file.
    """
    from skills.adp_run_automation import schedule_backend as sb

    sb.DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    path = sb.DOWNLOADS_DIR / f"Schedule-{datetime.date.today().isoformat()}.json"
    path.write_text(json.dumps(
        {
            "scraped_at_utc": datetime.datetime.utcnow().isoformat() + "Z",
            "store": store,
            "weeks": payloads,
        },
        indent=2,
    ))
    return path


def download_schedule(
    *,
    store: str = "palmetto",
    weeks: int = None,
    headed: bool = True,
    slow_mo_ms: int = 50,
    keep_open_on_error: bool = False,
) -> pathlib.Path:
    """Log in, open Team Schedule, scrape `weeks` of per-day totals, write JSON.

    Idempotency: if today's Schedule JSON is already on disk, skip the browser
    and return the cached path (avoids a duplicate ADP 2FA SMS on cron retries).
    """
    from skills.adp_run_automation import schedule_backend as sb

    weeks = weeks or sb.DEFAULT_WEEKS
    expected = sb.DOWNLOADS_DIR / f"Schedule-{datetime.date.today().isoformat()}.json"
    if is_fresh_download(expected, min_bytes=50):
        print(f"[adp_schedule] SKIP browser — fresh Schedule JSON already on disk: {expected}")
        return expected

    with launch_persistent(
        portal="adp",
        headed=headed,
        slow_mo_ms=slow_mo_ms,
        keep_open_on_error=keep_open_on_error,
    ) as (ctx, page):
        _ensure_logged_in(page, store=store)
        payloads = _schedule_within_session(page, weeks=weeks)
        return _write_schedule_json(payloads, store=store)


def download_payroll_liability(
    *,
    store: str = "palmetto",
    headed: bool = True,
    slow_mo_ms: int = 0,
    keep_open_on_error: bool = False,
) -> pathlib.Path:
    """Open Taxes → Tax reports → Payroll Liability; write JSON for BQ load.

    Best-effort channel — caller should not fail the nightly on liability errors.
    """
    out = DOWNLOADS_DIR / f"PayrollLiability-{datetime.date.today().isoformat()}.json"
    if is_fresh_download(out, min_bytes=100):
        print(f"[adp_liability] SKIP browser — fresh file: {out}")
        return out

    with launch_persistent(
        portal="adp",
        headed=headed,
        slow_mo_ms=slow_mo_ms,
        keep_open_on_error=keep_open_on_error,
    ) as (ctx, page):
        _ensure_logged_in(page, store=store)
        page.wait_for_timeout(2000)
        page.evaluate(
            """() => {
              const b = document.querySelector('[data-test-id="Taxes-btn"]');
              if (b) b.click();
            }"""
        )
        page.wait_for_timeout(2500)
        page.evaluate(
            """() => {
              const el = [...document.querySelectorAll('a,button,div')].find(
                e => /^Tax reports$/i.test((e.innerText || '').trim()));
              if (el) el.click();
            }"""
        )
        page.wait_for_timeout(3000)
        page.evaluate(
            """() => {
              const el = [...document.querySelectorAll('a,button')].find(
                e => /Payroll Liability/i.test((e.innerText || '').trim()));
              if (el) el.click();
            }"""
        )
        page.wait_for_timeout(4000)
        text = page.evaluate(
            "() => (document.body && document.body.innerText || '').replace(/\\s+/g, ' ').trim()"
        )
        payload = {
            "scraped_at_utc": datetime.datetime.utcnow().isoformat() + "Z",
            "store": store,
            "text": text,
            "reports": [text],
        }
        out.write_text(json.dumps(payload, indent=2))
        print(f"[adp_liability] wrote {out} ({len(text)} chars)")
        return out


# ── Bundle: one browser session, one login, both scrapes ───────────


def download_adp_bundle(
    *,
    store: str = "palmetto",
    target_date: Optional[datetime.date] = None,
    include_earnings: bool = True,
    include_schedule: bool = True,
    schedule_weeks: int = None,
    earnings_window_days: int = 90,
    earnings_start: Optional[datetime.date] = None,
    earnings_end: Optional[datetime.date] = None,
    earnings_custom_range: bool = False,
    headed: bool = True,
    slow_mo_ms: int = 50,
    keep_open_on_error: bool = False,
) -> dict:
    """Run BOTH ADP scrapes in a single browser session.

    Motivation: the standalone `download_timecard` and `download_earnings`
    functions each open their own context → each runs `_ensure_logged_in`
    → each can trigger ADP 2FA → operator gets TWO SMS OTPs per nightly
    run. This bundle opens ONE context, logs in ONCE, and runs both
    scrapes within that session, so the operator pays at most one OTP
    cost per refresh.

    Layered idempotency:
        - Layer A (file on disk): if today's XLSX is already on disk for a
          given component, that component is skipped and the cached path
          is returned without launching the browser.
        - Layer B (per-component markers): success of each scrape writes
          `~/.bhaga/state/run-<refresh_date>/{adp_timecard,adp_earnings}.done`
          (keyed by ``target_date`` = refresh_date, NOT today_ct — same
          rationale as daily_refresh._run_state_dir) so the orchestrator's
          per-step granularity is preserved even though it now invokes a
          single `adp_reports` step.

    Partial-success contract: if Timecard succeeds but Earnings fails (or
    vice versa), we DO NOT raise mid-flight. Both attempts run, the
    successful XLSX is returned in the result dict, and the caller can
    decide how to handle the partial failure (the orchestrator wraps this
    by raising AFTER inspecting `errors`, so the partial success is
    captured on disk + in markers before the exception propagates).

    Args:
        store: store profile name (Keychain entry key).
        target_date: date contained by the Timecard pay period to scrape.
            Passed through to `_timecard_within_session`. None = backfill.
        include_earnings: if False, only Timecard runs (orchestrator sets
            this off Mon/Tue per `_should_run_rates`).
        earnings_window_days: how far back the earnings scrape's "From"
            date should go when using the nightly "Last payroll" preset
            (default 90; ignored when earnings_custom_range=True).
        earnings_start: explicit window start for historical backfill
            (requires earnings_custom_range=True).
        earnings_end: explicit window end for historical backfill
            (requires earnings_custom_range=True).
        earnings_custom_range: if True, selects "Custom date range" in
            the ADP report builder and fills earnings_start/end + sets
            Employment status = All. Use for historical backfills only;
            nightly runs should keep False (Last payroll preset).

    Returns:
        {
            "timecard_xlsx": pathlib.Path | None,
            "earnings_xlsx": pathlib.Path | None,
            "errors":        {step_name: "ExcType: msg"},   # empty on full success
        }
    """
    today = datetime.date.today()
    tc_expected = DOWNLOADS_DIR / f"Timecard-{today.isoformat()}.xlsx"
    er_expected = DOWNLOADS_DIR / f"Earnings-and-Hours-V1-{today.isoformat()}.xlsx"
    sched_expected = DOWNLOADS_DIR / f"Schedule-{today.isoformat()}.json"

    # Layer A is target_date-aware (2026-05-23 fix): a file downloaded earlier
    # today for a DIFFERENT target_date does NOT count as fresh, because it
    # may not cover the current run's window. See _xlsx_fresh_for_target.
    tc_fresh = _xlsx_fresh_for_target(
        tc_expected, target_date=target_date, min_bytes=10_000,
    )
    er_fresh = (
        _xlsx_fresh_for_target(er_expected, target_date=target_date, min_bytes=5_000)
        if include_earnings else False
    )
    # Schedule is forward-looking (this week + next), not tied to target_date,
    # so a plain CT-today freshness check is sufficient.
    sched_fresh = (
        is_fresh_download(sched_expected, min_bytes=50) if include_schedule else False
    )

    result: dict = {
        "timecard_xlsx": tc_expected if tc_fresh else None,
        "earnings_xlsx": er_expected if (include_earnings and er_fresh) else None,
        "schedule_json": sched_expected if (include_schedule and sched_fresh) else None,
        "errors": {},
    }

    needs_timecard = not tc_fresh
    needs_earnings = include_earnings and not er_fresh
    needs_schedule = include_schedule and not sched_fresh

    if not needs_timecard and not needs_earnings and not needs_schedule:
        print("[adp_bundle] SKIP browser — Layer A: required ADP files already fresh on disk.")
        if tc_fresh:
            _mark_run_step_done(
                "adp_timecard", refresh_date=target_date,
                note="layer_a_skip (file fresh on disk)",
            )
        if include_earnings and er_fresh:
            _mark_run_step_done(
                "adp_earnings", refresh_date=target_date,
                note="layer_a_skip (file fresh on disk)",
            )
        if include_schedule and sched_fresh:
            _mark_run_step_done(
                "adp_schedule", refresh_date=target_date,
                note="layer_a_skip (file fresh on disk)",
            )
        return result

    print(f"[adp_bundle] needs_timecard={needs_timecard} needs_earnings={needs_earnings} "
          f"needs_schedule={needs_schedule}; opening single browser session (one login, one OTP cost).")

    with launch_persistent(
        portal="adp",
        headed=headed,
        slow_mo_ms=slow_mo_ms,
        keep_open_on_error=keep_open_on_error,
    ) as (ctx, page):
        _ensure_logged_in(page, store=store)
        dashboard_url = page.url
        print(f"[adp_bundle] dashboard_url={dashboard_url}")

        if needs_timecard:
            try:
                path = _timecard_within_session(
                    page, target_date=target_date, store=store
                )
                _write_target_meta(path, target_date)
                result["timecard_xlsx"] = path
                _mark_run_step_done(
                    "adp_timecard", refresh_date=target_date,
                    note=f"target_date={target_date.isoformat() if target_date else 'none'}",
                )
                print(f"[adp_bundle] timecard OK → {path}")
            except Exception as exc:  # noqa: BLE001
                result["errors"]["adp_timecard"] = f"{type(exc).__name__}: {exc}"
                print(f"[adp_bundle] timecard FAILED (continuing to earnings): "
                      f"{type(exc).__name__}: {exc}")
                # Per-component evidence — the launch_persistent global capture
                # only fires if the whole context with-block raises; here we
                # swallow to allow earnings to attempt.
                try:
                    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                    page.screenshot(
                        path=str(pathlib.Path.home() / ".bhaga" / "state" / "screenshots"
                                 / f"adp-bundle-timecard-fail-{ts}.png"),
                        full_page=True,
                    )
                except Exception:  # noqa: BLE001
                    pass

        if needs_earnings:
            try:
                # Navigate back to the dashboard URL (runpayrollmain.adp.com)
                # rather than LOGIN_URL (runpayroll.adp.com/enrollment.aspx). The login domain
                # doesn't share session cookies with the dashboard domain, so
                # hitting it triggers re-auth + a second OTP. Using the captured
                # dashboard_url keeps us on the same domain and preserves the
                # session. (2026-05-27 fix: single OTP per bundle)
                print(f"[adp_bundle] step=reset-page-before-earnings "
                      f"url={page.url}")
                page.goto(dashboard_url, wait_until="domcontentloaded", timeout=60_000)
                try:
                    page.wait_for_load_state("networkidle", timeout=30_000)
                except Exception:  # noqa: BLE001
                    pass
                if POST_LOGIN_URL_RE.search(page.url):
                    print(f"[adp_bundle] earnings prep: session alive "
                          f"(url={page.url})")
                else:
                    try:
                        page.wait_for_url(POST_LOGIN_URL_RE, timeout=10_000)
                        print(f"[adp_bundle] earnings prep: redirected to v2 "
                              f"(url={page.url})")
                    except Exception:  # noqa: BLE001
                        print(f"[adp_bundle] earnings: UNEXPECTED session lapse "
                              f"after timecard (url={page.url}, "
                              f"dashboard_url={dashboard_url}); re-running login")
                        _ensure_logged_in(page, store=store)
                        print(f"[adp_bundle] earnings: post-relogin "
                              f"url={page.url}")
                        page.wait_for_timeout(2_000)

                if earnings_custom_range and earnings_start and earnings_end:
                    window_start, window_end = earnings_start, earnings_end
                else:
                    window_end = target_date or today
                    window_start = window_end - datetime.timedelta(days=earnings_window_days)
                path = _earnings_within_session(
                    page, store=store, target_date=target_date,
                    start=window_start, end=window_end,
                    use_custom_range=earnings_custom_range,
                )
                _write_target_meta(path, target_date)
                result["earnings_xlsx"] = path
                _mark_run_step_done(
                    "adp_earnings", refresh_date=target_date,
                    note=f"window={window_start.isoformat()}..{window_end.isoformat()}",
                )
                print(f"[adp_bundle] earnings OK → {path}")
            except Exception as exc:  # noqa: BLE001
                result["errors"]["adp_earnings"] = f"{type(exc).__name__}: {exc}"
                print(f"[adp_bundle] earnings FAILED: {type(exc).__name__}: {exc}")
                try:
                    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                    page.screenshot(
                        path=str(pathlib.Path.home() / ".bhaga" / "state" / "screenshots"
                                 / f"adp-bundle-earnings-fail-{ts}.png"),
                        full_page=True,
                    )
                except Exception:  # noqa: BLE001
                    pass

        if needs_schedule:
            try:
                # Reuse the live session. Reset to the dashboard first (same
                # domain → no re-auth), then open Team Schedule and scrape.
                print(f"[adp_bundle] step=reset-page-before-schedule url={page.url}")
                page.goto(dashboard_url, wait_until="domcontentloaded", timeout=60_000)
                try:
                    page.wait_for_load_state("networkidle", timeout=30_000)
                except Exception:  # noqa: BLE001
                    pass
                if not POST_LOGIN_URL_RE.search(page.url):
                    try:
                        page.wait_for_url(POST_LOGIN_URL_RE, timeout=10_000)
                    except Exception:  # noqa: BLE001
                        print(f"[adp_bundle] schedule: UNEXPECTED session lapse "
                              f"(url={page.url}); re-running login")
                        _ensure_logged_in(page, store=store)
                payloads = _schedule_within_session(page, weeks=schedule_weeks)
                path = _write_schedule_json(payloads, store=store)
                result["schedule_json"] = path
                _mark_run_step_done(
                    "adp_schedule", refresh_date=target_date,
                    note=f"weeks={len(payloads)}",
                )
                print(f"[adp_bundle] schedule OK → {path}")
            except Exception as exc:  # noqa: BLE001
                result["errors"]["adp_schedule"] = f"{type(exc).__name__}: {exc}"
                print(f"[adp_bundle] schedule FAILED: {type(exc).__name__}: {exc}")
                try:
                    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
                    page.screenshot(
                        path=str(pathlib.Path.home() / ".bhaga" / "state" / "screenshots"
                                 / f"adp-bundle-schedule-fail-{ts}.png"),
                        full_page=True,
                    )
                except Exception:  # noqa: BLE001
                    pass

        # Employer burden (Payroll Liability) — best-effort, non-fatal.
        try:
            print(f"[adp_bundle] step=payroll-liability url={page.url}")
            page.goto(dashboard_url, wait_until="domcontentloaded", timeout=60_000)
            page.wait_for_timeout(2000)
            page.evaluate(
                """() => {
                  const b = document.querySelector('[data-test-id="Taxes-btn"]');
                  if (b) b.click();
                }"""
            )
            page.wait_for_timeout(2500)
            page.evaluate(
                """() => {
                  const el = [...document.querySelectorAll('a,button,div')].find(
                    e => /^Tax reports$/i.test((e.innerText || '').trim()));
                  if (el) el.click();
                }"""
            )
            page.wait_for_timeout(3000)
            page.evaluate(
                """() => {
                  const el = [...document.querySelectorAll('a,button')].find(
                    e => /Payroll Liability/i.test((e.innerText || '').trim()));
                  if (el) el.click();
                }"""
            )
            page.wait_for_timeout(4000)
            text = page.evaluate(
                "() => (document.body && document.body.innerText || '').replace(/\\s+/g, ' ').trim()"
            )
            liab_path = DOWNLOADS_DIR / f"PayrollLiability-{datetime.date.today().isoformat()}.json"
            liab_path.write_text(json.dumps({
                "scraped_at_utc": datetime.datetime.utcnow().isoformat() + "Z",
                "store": store,
                "text": text,
                "reports": [text],
            }, indent=2))
            result["liability_json"] = liab_path
            print(f"[adp_bundle] liability OK → {liab_path} ({len(text)} chars)")
        except Exception as exc:  # noqa: BLE001
            result["errors"]["adp_liability"] = f"{type(exc).__name__}: {exc}"
            print(f"[adp_bundle] liability FAILED: {type(exc).__name__}: {exc}")

        return result


# ── CLI ────────────────────────────────────────────────────────────


def main() -> int:
    cli = argparse.ArgumentParser(description=__doc__)
    cli.add_argument("scrape", choices=["timecard", "earnings", "schedule"],
                     help="Which scrape to run.")
    cli.add_argument("--store", default="palmetto")
    cli.add_argument("--headless", action="store_true",
                     help="Run without a visible browser window.")
    cli.add_argument("--keep-open", action="store_true",
                     help="On error, leave browser open for manual inspection.")
    cli.add_argument("--start", default=None,
                     help="(earnings only) Start date YYYY-MM-DD. Default: 90 days ago.")
    cli.add_argument("--end", default=None,
                     help="(earnings only) End date YYYY-MM-DD. Default: today.")
    args = cli.parse_args()

    if args.scrape == "timecard":
        path = download_timecard(
            store=args.store,
            headed=not args.headless,
            keep_open_on_error=args.keep_open,
        )
    elif args.scrape == "schedule":
        path = download_schedule(
            store=args.store,
            headed=not args.headless,
            keep_open_on_error=args.keep_open,
        )
    else:
        path = download_earnings(
            store=args.store,
            start_date=datetime.date.fromisoformat(args.start) if args.start else None,
            end_date=datetime.date.fromisoformat(args.end) if args.end else None,
            headed=not args.headless,
            keep_open_on_error=args.keep_open,
        )
    print(f"# Downloaded: {path} ({path.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
