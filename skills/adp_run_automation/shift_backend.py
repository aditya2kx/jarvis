#!/usr/bin/env python3
"""skills/adp_run_automation/shift_backend - Playwright-driven Timecard XLSX extractor.

Source: ADP RUN > Reports > Time reports > Timecard. The report renders in-page
(inside an iframe) for selected pay periods x all employees, then exports as a
flat .xlsx via an "Export To Excel" button. The xlsx 'Details' sheet is a clean
flat table (12 columns) with one row per punch (split-shift aware).

This is FAR cleaner than the alternative Time > Timecards modal which is
single-employee-at-a-time and would require iterating ~14 employees x ~5 pay
periods (~70 page loads) for a 2-month backfill.

Architecture mirrors skills/square_tips/transactions_backend.py:
    * `build_plan()` returns a deterministic Playwright playbook the AI agent
      executes via the `user-playwright` MCP. ADP login uses the
      `adp_palmetto_login` Keychain credential (handled by skills/credentials).
    * `parse_xlsx()` is pure-Python and unit-testable. Reads the 'Details' sheet
      and returns one record per punch.
    * `aggregate_by_day()` collapses per-punch records into per-(employee, date)
      rollups -- the shape the plan's `daily_shifts()` API specifies.
    * `daily_shifts()` is the public high-level entry: finds the most recent
      Timecard xlsx in extracted/downloads/, parses it, and returns the
      per-day rollups filtered to [start_date, end_date].

Critical gotchas (see selectors/timecards.json for full notes):
    * Open shifts (employee currently clocked in, no End Work) are EXCLUDED
      from the export. Schedule daily refresh AFTER all employees clock out
      (T-1 is safe).
    * ADP times are already in the shop's local TZ (unlike Square which is in
      account display TZ). No conversion needed.
    * Hours are in 'H:MM' format (HOURS:MINUTES, NOT decimal). 4:09 = 4.15h.
    * Some employees may appear under multiple name spellings (mid-period name
      edits in ADP). Apply employee_aliases for normalization. Real example
      from 2026-05-16 calibration: "Johnson Dolce" -> "Johnson Dolce J".
    * No `employee_id` column in the export. Until compensation_backend.py
      gives us ADP employee IDs, we use the normalized name as the id.
    * Pay periods are bi-weekly (for Palmetto). To cover a date window, select
      every pay period whose end >= window_start and start <= window_end.
      Over-coverage at the edges is filtered out in `daily_shifts()`.

Status (2026-05-16): proven from-scratch end-to-end on Palmetto Superfoods
account. 5 pay periods (Mar 9 - May 16, 2026), 14 employees, 606 punches.
"""

from __future__ import annotations

import datetime
import json
import os
import pathlib
import re
import subprocess
import sys
from typing import Optional

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))

from core.config_loader import project_dir
from skills.credentials import registry as cred_registry


_PROJECT = pathlib.Path(project_dir())
DOWNLOADS_DIR = _PROJECT / "extracted" / "downloads"
SELECTORS_PATH = _PROJECT / "skills" / "adp_run_automation" / "selectors" / "timecards.json"

LOGIN_URL = "https://runpayroll.adp.com"


# Open-shift exclusion: ADP RUN's Timecard Excel export omits ANY shift where
# the employee is currently clocked in (no End Work punch yet). The orchestrator
# (agents/bhaga/scripts/daily_refresh.py, built in M3) MUST schedule the daily
# refresh AFTER all employees have clocked out, otherwise those punches won't
# appear in that night's export and will only be picked up next time the
# shop-level window for that date is re-scraped.
#
# Safe pattern: always scrape T-1 (yesterday) at the daily refresh time, never
# T-0 (today). Per the Austin store profile, the schedule is 21:00 CT — well
# after Palmetto's typical 21:00 close — but a few stragglers can keep punching
# until ~21:30. Allowing 60 minutes of buffer keeps us safe.
#
# Implementation note for the M3 orchestrator:
#   * Read the store profile's `shop_close_local_time` (e.g. "21:00").
#   * The launchd plist should fire at shop_close_local_time + 60 minutes
#     (so e.g. 22:00 CT), and scrape T-1 (the previous calendar day).
#   * If a backfill run is requested for a date other than T-1, the data will
#     be complete because all that day's shifts have long since closed.
ORCHESTRATOR_SCRAPE_BUFFER_MINUTES_AFTER_SHOP_CLOSE = 60


# ── Credentials ───────────────────────────────────────────────────


def _credential_name(store: str) -> str:
    """Convention: 'adp_{store}_login' as a Keychain handle in skills/credentials."""
    return f"adp_{store.lower()}_login"


def get_credentials(store: str = "palmetto") -> dict:
    """Resolve ADP RUN login from Keychain. Returns {'username', 'password'}."""
    entry = cred_registry.lookup(_credential_name(store))
    if not entry:
        raise RuntimeError(
            f"No credential '{_credential_name(store)}' in registry. "
            f"Register it via skills/browser/collaborative.py or the credentials skill."
        )
    if entry.get("type") != "keychain":
        raise RuntimeError(
            f"Credential '{_credential_name(store)}' is type {entry.get('type')!r}, "
            f"expected 'keychain'."
        )
    result = subprocess.run(
        [
            "security", "find-generic-password",
            "-a", entry["account"],
            "-s", entry["service"],
            "-w",
        ],
        capture_output=True, text=True, timeout=5,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Keychain lookup failed for {entry['account']}@{entry['service']}: "
            f"{result.stderr.strip()}"
        )
    return {"username": entry["account"], "password": result.stdout.strip()}


# ── Selectors ─────────────────────────────────────────────────────


def selectors() -> dict:
    return json.loads(SELECTORS_PATH.read_text())


# ── Field parsing helpers ─────────────────────────────────────────


_DATE_PREFIXED_PATTERN = re.compile(r"^[A-Za-z]{3}\s+(\d{2}/\d{2}/\d{4})$")
_HHMM_PATTERN = re.compile(r"^(\d+):(\d{2})$")
_TIME_OF_DAY_PATTERN = re.compile(r"^(\d{1,2}):(\d{2})\s*(AM|PM)$", re.IGNORECASE)


def parse_punch_date(s: str) -> Optional[datetime.date]:
    """'Fri 05/15/2026' -> date(2026, 5, 15). Returns None on parse failure."""
    if not s:
        return None
    m = _DATE_PREFIXED_PATTERN.match(str(s).strip())
    if not m:
        return None
    try:
        return datetime.datetime.strptime(m.group(1), "%m/%d/%Y").date()
    except ValueError:
        return None


def parse_time_of_day(s: str) -> Optional[datetime.time]:
    """'1:32 PM' -> time(13, 32). Returns None on parse failure."""
    if not s:
        return None
    m = _TIME_OF_DAY_PATTERN.match(str(s).strip())
    if not m:
        return None
    hour, minute, meridiem = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if hour == 12:
        hour = 0
    if meridiem == "PM":
        hour += 12
    try:
        return datetime.time(hour, minute)
    except ValueError:
        return None


def parse_hhmm_to_decimal(s: str) -> float:
    """'4:09' -> 4.15. Empty/None -> 0.0.

    ADP stores hours as HOURS:MINUTES, NOT decimal. A 4:30 shift means
    4 hours 30 minutes = 4.5 decimal hours, NOT 4 hours 0.3 minutes. Bug
    waiting to happen if you treat the cell as a float directly.
    """
    if not s:
        return 0.0
    s = str(s).strip()
    if not s:
        return 0.0
    m = _HHMM_PATTERN.match(s)
    if not m:
        return 0.0
    return int(m.group(1)) + int(m.group(2)) / 60.0


def normalize_employee_name(name: str, aliases: Optional[dict] = None) -> str:
    """Apply alias map, trim whitespace, return canonical name.

    Example aliases (calibrated 2026-05-16):
        {"Johnson Dolce J": "Johnson Dolce"}

    Source-of-truth aliases live in
    agents/bhaga/knowledge-base/store-profiles/{store}.json under
    `employee_aliases`. Pass them in here per call.
    """
    if not name:
        return ""
    normalized = " ".join(str(name).split())  # collapse whitespace
    if aliases and normalized in aliases:
        return aliases[normalized]
    return normalized


# ── XLSX parsing ──────────────────────────────────────────────────


_DETAILS_COLUMNS = [
    "Employee Name",
    "Pay Period",
    "Date Range",
    "Total Paid Hours",
    "Date",
    "Start Work",
    "End Work",
    "Regular",
    "Overtime",
    "Doubletime",
    "Details",
    "Notes",
]


def parse_xlsx(
    xlsx_path: pathlib.Path,
    *,
    employee_aliases: Optional[dict] = None,
) -> list[dict]:
    """Parse the ADP Timecard XLSX 'Details' sheet into per-punch records.

    Each returned dict represents ONE punch pair (clock-in -> clock-out).
    Split shifts produce multiple records for the same (employee, date).

        {
            "employee_id": str,        # normalized name, until we wire ADP id
            "employee_name": str,      # normalized (alias-resolved)
            "raw_employee_name": str,  # original from the cell (for audit)
            "pay_period": str,         # 'YYYY-MM-DD to YYYY-MM-DD'
            "date": str,               # ISO 'YYYY-MM-DD' (shop-local)
            "in_time": str,            # 'HH:MM' 24h (shop-local)
            "out_time": str,           # 'HH:MM' 24h (shop-local)
            "regular_hours": float,    # decimal (NOT H:MM)
            "ot_hours": float,         # decimal
            "doubletime_hours": float, # decimal
            "punch_idx_in_day": int,   # 0 = first punch of the day, 1 = second, ...
            "raw_total_paid_hours": str,  # e.g. '54:53' (per-employee per-pp total)
        }

    Records sorted by (employee_name, date, in_time). Filters out malformed
    rows but does not enforce a date window -- pass through `daily_shifts()`
    for date filtering.

    Requires openpyxl. Install separately if not present (it's pulled in
    transitively by many data-science deps so most environments already have it).
    """
    import openpyxl

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Details" not in wb.sheetnames:
        raise ValueError(
            f"XLSX missing 'Details' sheet (found: {wb.sheetnames}). Make sure the "
            f"Timecard report was exported via 'Export To Excel' (not 'Print Report'); "
            f"the Print path produces a different shape."
        )

    ws = wb["Details"]
    header = [c.value for c in ws[1]]
    if header[: len(_DETAILS_COLUMNS)] != _DETAILS_COLUMNS:
        raise ValueError(
            f"Details sheet header mismatch. Expected {_DETAILS_COLUMNS}, "
            f"got {header[: len(_DETAILS_COLUMNS)]}. ADP may have changed the report layout."
        )

    records: list[dict] = []
    last_day_key: Optional[tuple] = None
    punch_idx = 0

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(header, raw_row))

        raw_name = row.get("Employee Name") or ""
        if not raw_name:
            continue

        # ADP includes "schedule" rows in the export — these show what the
        # employee was SCHEDULED to work (vs what they punched). They have
        # Details='Schedule', blank Date, blank Regular hours. They are not
        # actual worked time, so skip them. Calibrated 2026-05-16: 19 of 606
        # rows were schedule rows on Palmetto's report.
        if str(row.get("Details") or "").strip().lower() == "schedule":
            continue

        date_obj = parse_punch_date(row.get("Date"))
        if not date_obj:
            continue

        in_time_obj = parse_time_of_day(row.get("Start Work"))
        out_time_obj = parse_time_of_day(row.get("End Work"))
        if not in_time_obj or not out_time_obj:
            # Open shifts (no End Work) are normally excluded by ADP, but be
            # defensive: skip incomplete rows rather than emit a half-record.
            continue

        normalized_name = normalize_employee_name(raw_name, employee_aliases)

        day_key = (normalized_name, date_obj.isoformat())
        if day_key == last_day_key:
            punch_idx += 1
        else:
            punch_idx = 0
            last_day_key = day_key

        records.append({
            "employee_id": normalized_name,
            "employee_name": normalized_name,
            "raw_employee_name": str(raw_name).strip(),
            "pay_period": str(row.get("Pay Period") or "").strip(),
            "date": date_obj.isoformat(),
            "in_time": in_time_obj.strftime("%H:%M"),
            "out_time": out_time_obj.strftime("%H:%M"),
            "regular_hours": parse_hhmm_to_decimal(row.get("Regular")),
            "ot_hours": parse_hhmm_to_decimal(row.get("Overtime")),
            "doubletime_hours": parse_hhmm_to_decimal(row.get("Doubletime")),
            "punch_idx_in_day": punch_idx,
            "raw_total_paid_hours": str(row.get("Total Paid Hours") or "").strip(),
        })

    records.sort(key=lambda r: (r["employee_name"], r["date"], r["in_time"]))
    return records


# ── Aggregation ───────────────────────────────────────────────────


def aggregate_by_day(punches: list[dict]) -> list[dict]:
    """Collapse per-punch records into one record per (employee, date).

    Matches the public API shape promised in the bhaga-daily-refresh plan:

        {
            "date": "YYYY-MM-DD",
            "employee_id": str,
            "employee_name": str,
            "in_time": str,           # earliest punch-in of the day (HH:MM)
            "out_time": str,          # latest punch-out of the day (HH:MM)
            "regular_hours": float,   # sum across all punches that day
            "ot_hours": float,        # sum
            "doubletime_hours": float,# sum
            "total_hours": float,     # regular + ot + doubletime
            "punch_count": int,       # 2 for typical split shift, 1 for no-break
        }
    """
    by_key: dict[tuple, dict] = {}
    for p in punches:
        key = (p["employee_id"], p["date"])
        if key not in by_key:
            by_key[key] = {
                "date": p["date"],
                "employee_id": p["employee_id"],
                "employee_name": p["employee_name"],
                "in_time": p["in_time"],
                "out_time": p["out_time"],
                "regular_hours": 0.0,
                "ot_hours": 0.0,
                "doubletime_hours": 0.0,
                "total_hours": 0.0,
                "punch_count": 0,
            }
        bucket = by_key[key]
        bucket["in_time"] = min(bucket["in_time"], p["in_time"])
        bucket["out_time"] = max(bucket["out_time"], p["out_time"])
        bucket["regular_hours"] += p["regular_hours"]
        bucket["ot_hours"] += p["ot_hours"]
        bucket["doubletime_hours"] += p["doubletime_hours"]
        bucket["total_hours"] = (
            bucket["regular_hours"] + bucket["ot_hours"] + bucket["doubletime_hours"]
        )
        bucket["punch_count"] += 1

    out = list(by_key.values())
    out.sort(key=lambda r: (r["date"], r["employee_name"]))
    return out


# ── Playwright playbook ───────────────────────────────────────────


def build_plan(
    start_date: datetime.date,
    end_date: datetime.date,
    *,
    store: str = "palmetto",
    creds: Optional[dict] = None,
) -> dict:
    """Deterministic Playwright playbook for the AI to execute via user-playwright.

    Flow:
        1. Navigate to LOGIN_URL (ADP may recognize existing session).
        2. Enter password (Keychain-resolved at runtime, NOT embedded in plan).
        3. Click Reports in top nav. Find the 'Single reports' modal.
        4. Expand the 'Time reports' accordion. Click 'Timecard'.
        5. Inside the iframe[name='mdfTimeFrame']:
           - Open Pay Period multi-select; tick every pay period that overlaps
             [start_date, end_date]. (Over-coverage at edges is fine; trimmed
             in daily_shifts().)
           - Format = 'Continuous'
           - Click 'Apply Changes' and wait for render.
        6. Click Export To Excel (#report-excel-button via iframe DOM eval).
        7. The browser MCP downloads to extracted/downloads/Timecard-.xlsx.

    The plan does NOT enumerate pay periods up front (we don't know what's
    available without calibrating). Instead, the AI reads the pay-period
    listbox, filters to those overlapping the requested window, and ticks them.
    """
    sels = selectors()
    creds = creds or get_credentials(store)

    return {
        "store": store.lower(),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "creds_username": creds["username"],  # password NOT embedded
        "captures": {
            "xlsx_path": None,
            "pay_periods_selected": [],
            "errors": [],
        },
        "steps": [
            {
                "id": "navigate_login",
                "action": "browser_navigate",
                "args": {"url": LOGIN_URL},
                "description": (
                    "Open ADP RUN login. If prior session cookies exist, the page shows "
                    "'Hello, {Name}.' and prompts only for password. If fresh session, the "
                    "username-first flow needs separate calibration (NOT yet done -- M3 will "
                    "seed via a persistent Playwright profile)."
                ),
                "postcondition": "URL contains 'online.adp.com/signin' OR 'runpayrollmain.adp.com'.",
            },
            {
                "id": "login_password",
                "action": "browser_type",
                "selectors_hint": "iframe-less. getByRole('textbox', { name: 'Password' })",
                "value_ref": "creds.password",
                "args": {"submit": True},
                "description": (
                    "Type password and submit via Enter (the Sign in button starts disabled). "
                    "Password resolved at runtime via Keychain handle "
                    f"'{sels['login']['credential_handle']}'."
                ),
                "postcondition": "URL matches /runpayrollmain.adp.com/.*/v2/",
            },
            {
                "id": "wait_for_dashboard",
                "action": "browser_wait_for",
                "args": {"time": 4},
                "description": "Allow SPA to mount.",
            },
            {
                "id": "open_reports",
                "action": "browser_click",
                "selectors_hint": "Top nav button: data-test-id='Reports-btn' OR getByRole('button',{name:'Reports'})",
                "description": "Opens the Reports landing or Single reports modal.",
            },
            {
                "id": "wait_for_reports_modal",
                "action": "browser_wait_for",
                "args": {"time": 3},
                "description": "Reports modal renders the categories accordion.",
            },
            {
                "id": "expand_time_reports",
                "action": "browser_click",
                "selectors_hint": "Button with text 'Time reports' (h4 + show-content toggle)",
                "description": "Expand the Time reports category in the Single reports modal.",
            },
            {
                "id": "open_timecard_report",
                "action": "browser_click",
                "selectors_hint": "Link in the expanded list with text 'Timecard' (NOT 'Timecard Edit Audit')",
                "description": (
                    "Opens the Timecard report modal. Contents inside iframe[name='mdfTimeFrame']. "
                    "Filters render first; the report body renders after Apply Changes."
                ),
            },
            {
                "id": "wait_for_iframe",
                "action": "browser_wait_for",
                "args": {"time": 5},
                "description": "Allow the mdfTimeFrame iframe to load. Filters become reffable.",
            },
            {
                "id": "open_pay_period_multiselect",
                "action": "browser_click",
                "selectors_hint": (
                    "iframe('mdfTimeFrame').getByRole('combobox', { name: /Pay Period.*Select All/i })"
                ),
                "description": "Open the multi-select. Listbox renders below with each pay period as an option.",
            },
            {
                "id": "select_overlapping_pay_periods",
                "action": "python_then_browser_click",
                "description": (
                    f"For each option text matching /(\\d{{4}}-\\d{{2}}-\\d{{2}}) - (\\d{{4}}-\\d{{2}}-\\d{{2}})/, "
                    f"parse start/end dates and click the option if it overlaps "
                    f"[{start_date.isoformat()}, {end_date.isoformat()}]. "
                    "Push each selected period text to captures.pay_periods_selected. "
                    "Pay period options live inside iframe('mdfTimeFrame') role='option' with name regex above."
                ),
            },
            {
                "id": "close_dropdown",
                "action": "browser_press_key",
                "args": {"key": "Escape"},
                "description": "Close the listbox so the report body becomes clickable.",
            },
            {
                "id": "set_format_continuous",
                "action": "browser_click_then_option",
                "click_hint": "iframe('mdfTimeFrame').getByRole('combobox', { name: /Format/ })",
                "option_hint": "iframe('mdfTimeFrame').getByRole('option', { name: 'Continuous' })",
                "description": (
                    "Switch Format to 'Continuous'. Affects HTML layout only -- the Excel export is "
                    "format-agnostic -- but Continuous keeps the inline render simpler if we ever fall back to DOM scraping."
                ),
                "skip_if": "Combobox already shows 'Continuous'.",
            },
            {
                "id": "apply_changes",
                "action": "browser_click",
                "selectors_hint": "iframe('mdfTimeFrame').getByRole('button', { name: 'Apply Changes' })",
                "description": "Triggers report re-render. Takes 2-6 seconds for ~5 pay periods x ~14 employees.",
            },
            {
                "id": "wait_for_report_body",
                "action": "browser_wait_for",
                "args": {"time": 8},
                "description": (
                    "Wait for report body to render fully. Confirmation: the Export To Excel button "
                    "(id='report-excel-button' inside the iframe) becomes available."
                ),
            },
            {
                "id": "click_export_to_excel",
                "action": "browser_evaluate",
                "function": (
                    "() => { "
                    "const f = document.querySelector('iframe[name=\"mdfTimeFrame\"]'); "
                    "const btn = f.contentDocument.getElementById('report-excel-button'); "
                    "if (!btn) throw new Error('export button not found'); "
                    "btn.click(); return { clicked: true }; "
                    "}"
                ),
                "description": (
                    "Click via direct DOM eval -- the button is a custom SDF-BUTTON element with stable "
                    "id but role-based selectors don't always reach into shadow DOM cleanly. Triggers "
                    "an immediate download (no async generation). File lands at "
                    "extracted/downloads/Timecard-.xlsx (trailing space in source name gets normalized)."
                ),
            },
            {
                "id": "wait_for_download",
                "action": "browser_wait_for",
                "args": {"time": 5},
                "description": "Let the file flush to disk before parse_xlsx() reads it.",
            },
            {
                "id": "parse_and_validate",
                "action": "python",
                "description": (
                    "punches = shift_backend.parse_xlsx(captures.xlsx_path, employee_aliases=...). "
                    "daily = shift_backend.aggregate_by_day(punches). "
                    "Validate: every captures.pay_periods_selected appears in {p['pay_period'] for p in punches}; "
                    "if any selected pay period is missing entirely from the data, alert via slack (likely "
                    "filter didn't apply)."
                ),
            },
        ],
    }


# ── Public entry point ────────────────────────────────────────────


def daily_shifts(
    start_date: datetime.date,
    end_date: datetime.date,
    *,
    store: str = "palmetto",
    employee_aliases: Optional[dict] = None,
) -> list[dict]:
    """High-level entry: parse the most recent Timecard XLSX and return per-day rollups.

    Filters to [start_date, end_date] inclusive (trims the over-coverage that
    happens when pay-period boundaries don't align with the window).

    Returns the shape promised in the bhaga-daily-refresh plan:
        [{date, employee_id, employee_name, in_time, out_time, regular_hours,
          ot_hours, doubletime_hours, total_hours, punch_count}, ...]
    """
    candidates = sorted(
        DOWNLOADS_DIR.glob("Timecard*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"No Timecard*.xlsx found in {DOWNLOADS_DIR}. Run build_plan() and have the "
            f"AI drive Playwright to populate."
        )
    punches = parse_xlsx(candidates[0], employee_aliases=employee_aliases)
    daily = aggregate_by_day(punches)
    start_iso, end_iso = start_date.isoformat(), end_date.isoformat()
    return [r for r in daily if start_iso <= r["date"] <= end_iso]


def raw_punches(
    start_date: datetime.date,
    end_date: datetime.date,
    *,
    store: str = "palmetto",
    employee_aliases: Optional[dict] = None,
) -> list[dict]:
    """Like daily_shifts() but returns per-punch records (one row per punch pair)."""
    candidates = sorted(
        DOWNLOADS_DIR.glob("Timecard*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"No Timecard*.xlsx found in {DOWNLOADS_DIR}."
        )
    punches = parse_xlsx(candidates[0], employee_aliases=employee_aliases)
    start_iso, end_iso = start_date.isoformat(), end_date.isoformat()
    return [p for p in punches if start_iso <= p["date"] <= end_iso]


# ── CLI ───────────────────────────────────────────────────────────


if __name__ == "__main__":
    import argparse

    cli = argparse.ArgumentParser(description="ADP RUN Timecard report extractor")
    sub = cli.add_subparsers(dest="cmd")

    p_parse = sub.add_parser("parse", help="Parse an existing Timecard XLSX.")
    p_parse.add_argument("xlsx_path")
    p_parse.add_argument(
        "--aliases", default=None,
        help="JSON file with {raw_name: canonical_name} alias map.",
    )
    p_parse.add_argument(
        "--rollup", action="store_true",
        help="Aggregate to per-day per-employee instead of per-punch.",
    )

    p_daily = sub.add_parser("daily", help="Parse most recent download and filter to a date window.")
    p_daily.add_argument("--start", required=True, help="YYYY-MM-DD")
    p_daily.add_argument("--end", required=True, help="YYYY-MM-DD")
    p_daily.add_argument("--store", default="palmetto")
    p_daily.add_argument("--aliases", default=None)

    p_plan = sub.add_parser("plan", help="Print the Playwright playbook for a date range.")
    p_plan.add_argument("--start", required=True, help="YYYY-MM-DD")
    p_plan.add_argument("--end", required=True, help="YYYY-MM-DD")
    p_plan.add_argument("--store", default="palmetto")

    p_creds = sub.add_parser("verify-creds", help="Check Keychain access.")
    p_creds.add_argument("--store", default="palmetto")

    args = cli.parse_args()
    aliases = json.loads(pathlib.Path(args.aliases).read_text()) if getattr(args, "aliases", None) else None

    if args.cmd == "parse":
        punches = parse_xlsx(pathlib.Path(args.xlsx_path), employee_aliases=aliases)
        if args.rollup:
            print(json.dumps(aggregate_by_day(punches), indent=2))
        else:
            print(json.dumps(punches, indent=2))
    elif args.cmd == "daily":
        records = daily_shifts(
            datetime.date.fromisoformat(args.start),
            datetime.date.fromisoformat(args.end),
            store=args.store,
            employee_aliases=aliases,
        )
        print(json.dumps(records, indent=2))
    elif args.cmd == "plan":
        plan = build_plan(
            datetime.date.fromisoformat(args.start),
            datetime.date.fromisoformat(args.end),
            store=args.store,
        )
        print(json.dumps(plan, indent=2))
    elif args.cmd == "verify-creds":
        creds = get_credentials(args.store)
        print(json.dumps({
            "store": args.store,
            "username": creds["username"],
            "password_length": len(creds["password"]),
            "password_preview": creds["password"][:2] + "***" + creds["password"][-2:],
        }, indent=2))
    else:
        cli.print_help()
